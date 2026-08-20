import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def populate_audi_clean_desc_remarks():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)
    
    if "Audi" not in wb.sheetnames:
        ws = wb.create_sheet(title="Audi")
    else:
        ws = wb["Audi"]

    # Clear sheet content to rebuild clean headers and data
    ws.delete_rows(1, ws.max_row + 1)
    ws.views.sheetView[0].showGridLines = True

    # Styling Setup
    font_family = "TH Sarabun New"
    
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font = Font(name=font_family, size=14, color="333333")
    bold_font = Font(name=font_family, size=14, bold=True, color="0F4C81")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    headers = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "คำอธิบาย (Description)",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_idx in range(1, 12):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Model Group Definitions
    G1 = ["A1", "A3", "Q2", "Q3"]
    G2 = ["A4", "A5", "A6", "Q5"]
    G3 = ["A7", "A8", "Q7", "Q8", "R8", "TT", "TTs"]
    G3_NO_R8 = ["A7", "A8", "Q7", "Q8", "TT", "TTs"]

    # Items Data Definition
    RAW_ITEMS_DATA = [
        # Page 1 Items
        ("กันชนหน้า", "Center", [
            (G1, 9000, 8000, 10000, 12000),
            (G2, 10500, 9500, 11500, 13500),
            (G3, 12000, 11000, 13000, 15000),
        ]),
        ("ฝากระโปรงหน้า", "Center", [
            (G1, 10500, 9500, 11500, 13500),
            (G2, 12000, 11000, 13000, 15000),
            (G3, 13500, 12500, 14500, 16500),
        ]),
        ("บังโคลนหน้า", "L/R", [
            (G1, 8500, 7500, 9500, 11500),
            (G2, 10000, 9000, 11000, 13000),
            (G3, 11500, 10500, 12500, 14500),
        ]),
        ("ประตูหน้า", "L/R", [
            (G1, 12000, 8500, 10500, 12500),
            (G2, 14000, 10000, 13500, 14000),
            (G3, 16000, 11500, 13500, 15500),
        ]),
        ("ประตูหลัง", "L/R", [
            (G1, 12000, 8500, 10500, 12500),
            (G2, 14000, 10000, 12000, 14000),
            (G3, 16000, 11500, 13500, 15500),
        ]),
        ("บังโคลนหลัง", "L/R", [
            (G1, 19000, 9500, 11500, 13500),
            (G2, 22000, 11000, 13000, 15000),
            (G3, 24500, 12500, 14500, 16500),
        ]),
        ("หลังคา", "Center", [
            (G1, 18500, 11500, 13500, 15500),
            (G2, 20000, 13000, 15000, 17000),
            (G3, 21500, 14500, 16500, 18500),
        ]),
        ("ฝากระโปรงหลัง", "Center", [
            (G1, 9500, 8500, 10500, 12500),
            (G2, 11000, 10000, 12000, 14000),
            (G3, 12500, 11500, 13500, 15500),
        ]),

        # Page 2 Items
        ("กันชนหลัง", "Center", [
            (G1, 9000, 8000, 10000, 12000),
            (G2, 10500, 9500, 11500, 13500),
            (G3, 12000, 11000, 13000, 15000),
        ]),
        ("กระจกมองข้าง", "L/R", [
            (G1, 2500, 2000, None, None),
            (G2, 2500, 2200, None, None),
            (G3, 3000, 2500, None, None),
        ]),
        ("สเกิร์ตบันได", "L/R", [
            (G1, 6000, 6000, None, None),
            (G2, 6500, 6000, None, None),
            (G3, 7000, 6000, None, None),
        ]),
        ("ล้อแม็ค", "L/R", [
            (G1, None, 3500, None, None),
            (G2, None, 4000, None, None),
            (G3, None, 4500, None, None),
        ]),
        ("ถอด-ใส่กระจกบังลมหน้า", "Center", [
            (G1, None, 5500, None, None),
            (G2, None, 5500, None, None),
            (G3_NO_R8, None, 5500, None, None),
            (["R8"], None, 7000, None, None),
        ]),
        ("ถอด-ใส่กระจกบังลมหลัง", "Center", [
            (G1, None, 5500, None, None),
            (G2, None, 5500, None, None),
            (G3_NO_R8, None, 5500, None, None),
            (["R8"], None, 7000, None, None),
        ]),
        ("มือจับ", "L/R", [
            (G1, None, 2000, None, None),
            (G2, None, 2000, None, None),
            (G3_NO_R8, None, 2000, None, None),
            (["R8"], None, 2000, None, None),
        ]),
    ]

    row_count = 1
    current_excel_row = 2

    for subject, position_lr, price_groups in RAW_ITEMS_DATA:
        for model_group, replace_p, s_p, m_p, l_p in price_groups:
            for model_name in model_group:
                row_values = [
                    row_count,
                    "Audi",
                    model_name,
                    subject,
                    position_lr,
                    None, # Column F: Description (Blank)
                    replace_p,
                    s_p,
                    m_p,
                    l_p,
                    None  # Column K: Remark (Blank)
                ]
                ws.append(row_values)
                
                # Format cell
                fill = zebra_fill if current_excel_row % 2 == 0 else white_fill
                for col_idx in range(1, 12):
                    cell = ws.cell(row=current_excel_row, column=col_idx)
                    cell.fill = fill
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if col_idx in [1, 2, 3, 5]: # No, Brand, Model, Position
                        cell.alignment = align_center
                        if col_idx == 3:
                            cell.font = bold_font
                    elif col_idx in [7, 8, 9, 10]: # Prices
                        cell.alignment = align_right
                        if cell.value is not None:
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = align_left

                ws.row_dimensions[current_excel_row].height = 22
                row_count += 1
                current_excel_row += 1

    # Auto adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws.column_dimensions['D'].width = 28 # Subject
    ws.column_dimensions['E'].width = 16 # Position (L/R)
    ws.column_dimensions['F'].width = 22 # Description (Blank)
    ws.column_dimensions['K'].width = 22 # Remark (Blank)

    wb.save(excel_path)
    print(f"SUCCESS: Cleared description and remarks in Audi sheet at {excel_path}")

if __name__ == "__main__":
    populate_audi_clean_desc_remarks()
