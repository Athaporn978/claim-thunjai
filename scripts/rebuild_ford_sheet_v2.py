#!/usr/bin/env python3
"""
rebuild_ford_sheet_v2.py
Clears Sheet FORD and rebuilds from scratch using all 4 Ford PDF files.
- Headers: copied from Audi sheet (12 columns)
- Data: extracted from price tables in each PDF
- Thai text: cleaned using targeted replacement (no NFC normalization)
"""

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

EXCEL_PATH = '/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx'
FORD_DIR = '/Users/arthur/Documents/Customer/H Tech/Labor Motor/FORD/'

# ============================================================
# Thai Text Cleaning (targeted replacement only, NO NFC)
# ============================================================
THAI_FIXES = [
    # broken combining chars
    ('กั ้น', 'กั้น'),
    ('ตั ้ง', 'ตั้ง'),
    ('ติดตั ้ง', 'ติดตั้ง'),
    ('ตัั ้ง', 'ตั้ง'),
    # specific known issues
    ('เพ็ องท้าย', 'เฟืองท้าย'),
    ('เพืองท้าย', 'เฟืองท้าย'),
    ('กระบะทั ้งลูก', 'กระบะทั้งลูก'),
    ('ถีั ้เพลิง', 'ถังเพลิง'),
    ('จุดเข็มขัดนิรภัย', 'ชุดเข็มขัดนิรภัย'),
    ('ฝาปิด ที่ฉีด', 'ฝาปิดที่ฉีด'),
    ('ทโู ทน', 'ทูโทน'),
    ('ด า', 'ดำ'),
    ('ด ำ', 'ดำ'),
    ('โครเมยี ม', 'โครเมียม'),
    ('ตัวลอ๊ ค', 'ตัวล็อก'),
    ('ซมุ้', 'ซุ้ม'),
    ('แชสซสี ์', 'แชสซีส์'),
    ('แชสซี ์', 'แชสซีส์'),
    # broken vowels/consonants from CID font
    ('หนา้', 'หน้า'),
    ('หลงั ', 'หลัง'),
    ('หลงั', 'หลัง'),
    ('ซา้ ย', 'ซ้าย'),
    ('ซา้ย', 'ซ้าย'),
    ('ขวา ', 'ขวา'),
    ('ตอ่ ', 'ต่อ'),
    ('ตอ่', 'ต่อ'),
    ('ขา้ ง', 'ข้าง'),
    ('ขา้ง', 'ข้าง'),
    ('ดา้ น', 'ด้าน'),
    ('ดา้น', 'ด้าน'),
    ('ลา้ ง', 'ล้าง'),
    ('กา้ น', 'ก้าน'),
    ('บา้ ง', 'บ้าง'),
    ('ท า', 'ทำ'),
    ('ท ำ', 'ทำ'),
    ('ทา สี', 'ทำสี'),
    ('ทา สี', 'ทำสี'),
    ('สา หรับ', 'สำหรับ'),
    ('สา หรับ', 'สำหรับ'),
    ('ยา้ ย', 'ย้าย'),
    ('ลอ้ ', 'ล้อ'),
    ('ลอ้', 'ล้อ'),
    ('รอ้ ย', 'ร้อย'),
    ('ฟอรด์', 'ฟอร์ด'),
    ('ฟอรด์ ', 'ฟอร์ด '),
    ('ฟอรด์\n', 'ฟอร์ด\n'),
    ('ถว่ ง', 'ถ่วง'),
    ('ถว่ง', 'ถ่วง'),
    ('ตงั้ ', 'ตั้ง'),
    ('ตงั้', 'ตั้ง'),
    ('คา่ ', 'ค่า'),
    ('คา่', 'ค่า'),
    ('ทัง้ ', 'ทั้ง'),
    ('ทัง้', 'ทั้ง'),
    ('ทงั้ ', 'ทั้ง'),
    ('ทงั้', 'ทั้ง'),
    ('ขนา้ ง', 'ข้าง'),
    ('ชดุ ', 'ชุด'),
    ('ชดุ', 'ชุด'),
    ('เปลยี่ น', 'เปลี่ยน'),
    ('เปลยี่น', 'เปลี่ยน'),
    ('บนั ได', 'บันได'),
    ('บนั ได', 'บันได'),
    ('แร็ค', 'แร็ค'),
    ('ผา้ ', 'ผ้า'),
    ('ผา้', 'ผ้า'),
    ('บหุ ลงั คา', 'บุหลังคา'),
    ('บหุ ลงั คา', 'บุหลังคา'),
    ('หลงั คา', 'หลังคา'),
    ('หลงั คา', 'หลังคา'),
    ('รวมคา่ ', 'รวมค่า'),
    ('รวมคา่', 'รวมค่า'),
    ('ถอดประกอบ', 'ถอดประกอบ'),
    ('ฝาปิดหลู าก', 'ฝาปิดฝูลาก'),
    ('รังผงึ้ ', 'รังผึ้ง'),
    ('รังผงึ้', 'รังผึ้ง'),
    ('ศนู ย์', 'ศูนย์'),
    ('ศนู ย์', 'ศูนย์'),
    ('ทสี่ ี', 'ที่สี'),
    ('ทา สแี ', 'ทำสีแ'),
    ('ทา สี', 'ทำสี'),
    ('กระจกมองขา้ ง', 'กระจกมองข้าง'),
    ('กระจกมองขา้ง', 'กระจกมองข้าง'),
    ('ไมร่ วม', 'ไม่รวม'),
    ('ไมร่วม', 'ไม่รวม'),
    ('นวิ้ ', 'นิ้ว'),
    ('นวิ้', 'นิ้ว'),
    ('ตอ่ วง', 'ต่อวง'),
    ('บานพับ', 'บานพับ'),
    ('ประตหู ', 'ประตู'),
    ('กนั ชน', 'กันชน'),
    ('แถบซลี ', 'แถบซีล'),
    ('แถบซลี', 'แถบซีล'),
    ('ยาง ซลี ', 'ยางซีล'),
    ('ซลี ', 'ซีล'),
    ('ซลี', 'ซีล'),
    ('เสาหลงั คา', 'เสาหลังคา'),
    ('คานเหล็กตอ่ หวั ', 'คานเหล็กต่อหัว'),
    ('คัชช ีตอ่ ขา้ ง', 'คัชชีต่อข้าง'),
    ('คาน', 'คาน'),
    ('พนังั ', 'พนัง'),
    ('พนัง', 'พนัง'),
    ('สลี อ้ ', 'สีล้อ'),
    ('สลี อ้', 'สีล้อ'),
    ('ทา สลี อ้ ', 'ทำสีล้อ'),
    ('ทา สลี อ้', 'ทำสีล้อ'),
    ('ขอบ18-20', 'ขอบ 18-20'),
    ('*ราคาสทุ ธ ิไมม่ สี ว่ นลด', '*ราคาสุทธิไม่มีส่วนลด'),
    ('*ราคาสทุ ธิไมม่ สีว่ นลด', '*ราคาสุทธิไม่มีส่วนลด'),
    ('**ราคาสา หรับรถปกต ิ', '**ราคาสำหรับรถปกติ'),
    ('**ราคาสำหรับรถปกต ิ', '**ราคาสำหรับรถปกติ'),
    ('ไมม่ สี ว่ นลด', 'ไม่มีส่วนลด'),
    ('ยางมาตรฐานฟอรด์', 'ยางมาตรฐานฟอร์ด'),
    ('Wildtrak (ทโู ทน) ด า+โครเมยี ม', 'Wildtrak (ทูโทน) ดำ+โครเมียม'),
    ('Wildtrak (ทูโทน) ด า+โครเมียม', 'Wildtrak (ทูโทน) ดำ+โครเมียม'),
]

def clean_thai(text: str) -> str:
    if not text:
        return ''
    text = str(text).strip()
    for wrong, right in THAI_FIXES:
        text = text.replace(wrong, right)
    # Remove excessive whitespace
    text = re.sub(r'  +', ' ', text)
    text = text.strip()
    return text

def parse_price(val):
    if val is None:
        return None
    s = str(val).strip().replace(',', '').replace(' ', '').replace('\n', '')
    try:
        return float(s)
    except ValueError:
        return None

# ============================================================
# Sub-model detection from page title text
# ============================================================
def detect_submodel(page_text: str, is_suv: bool) -> str:
    text = page_text or ''
    if is_suv:
        return 'SUV'
    if 'Single Cab' in text:
        return 'Single Cab'
    if 'Open Cab' in text:
        return 'Open Cab'
    if 'Double Cab' in text:
        return 'Double Cab'
    # Front/Center Rear without cab = Full (whole body) - use the section type
    if 'Front' in text and 'Truck' not in text.split('Front')[0][-20:]:
        return 'Full'
    return 'Full'

def detect_position(thai_text: str) -> str:
    """Detect L/R/L-R from Thai text"""
    t = thai_text or ''
    if 'ซ้าย-ขวา' in t or 'ซ้าย - ขวา' in t or 'L-R' in t or 'L/R' in t:
        return 'L/R'
    if 'ซ้าย' in t and 'ขวา' not in t:
        return 'L'
    if 'ขวา' in t and 'ซ้าย' not in t:
        return 'R'
    return ''

# ============================================================
# Extract rows from a price-table page
# ============================================================
def extract_rows_from_table(table, submodel: str, model: str) -> list[dict]:
    """
    Given a pdfplumber table (list of rows), extract data rows.
    Returns list of dicts with keys: submodel, subject, position, replace, s, m, l
    """
    rows_out = []
    # Detect column layout by finding header row
    header_row_idx = None
    for idx, row in enumerate(table):
        # Header row has 'รายการ' or 'Description'
        row_str = ' '.join([str(c) for c in row if c])
        if 'รายการ' in row_str or 'Description' in row_str:
            header_row_idx = idx
            break

    if header_row_idx is None:
        return []

    header = table[header_row_idx]
    ncols = len(header)

    # Map column indices
    # Possible formats:
    # Format A (9 cols): No | JobL | JobR | Description | รายการ | ประกอบพ่นสี | เบา | กลาง | หนัก
    # Format B (10 cols): No | JobL | JobC | JobR | Description | รายการ | ประกอบพ่นสี | เบา | กลาง | หนัก
    # Format C (3 cols): No | รายการ | ราคา (labor list)
    
    col_no = col_thai = col_replace = col_s = col_m = col_l = None

    if ncols == 3:
        # Labor list format
        col_no, col_thai, col_replace = 0, 1, 2
        col_s = col_m = col_l = None
    elif ncols >= 9:
        col_no = 0
        # Find รายการ column
        for ci, h in enumerate(header):
            if h and 'รายการ' in str(h):
                col_thai = ci
            if h and ('ประกอบ' in str(h) or 'พ่นสี' in str(h) or 'เปลี่ยน' in str(h)):
                col_replace = ci
        
        # prices: last 3 columns = เบา, กลาง, หนัก
        # but ประกอบพ่นสี is actually col H (Replace)
        # layout: ... | ประกอบพ่นสี | เบา | กลาง | หนัก
        # so: replace = ncols-4, s = ncols-3, m = ncols-2, l = ncols-1
        col_replace = ncols - 4
        col_s = ncols - 3
        col_m = ncols - 2
        col_l = ncols - 1

        if col_thai is None:
            # fallback: รายการ is col before ประกอบพ่นสี
            col_thai = col_replace - 1

    for row in table[header_row_idx + 1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        
        # Skip rows that are sub-headers (contain only text, no prices)
        row_str = ' '.join([str(c) for c in row if c])
        
        # Get Thai subject
        if col_thai is not None and col_thai < len(row):
            thai_raw = str(row[col_thai] or '').strip()
        else:
            continue

        if not thai_raw or thai_raw == 'None':
            continue

        # Skip pure header/title rows
        if any(kw in thai_raw for kw in ['รายการ', 'Description', 'ประกอบ', 'เบา', 'กลาง', 'หนัก']):
            if col_replace is not None and col_replace < len(row):
                price_check = parse_price(row[col_replace])
                if price_check is None:
                    continue

        thai_clean = clean_thai(thai_raw)
        position = detect_position(thai_clean)
        
        # Remove position text from subject
        for pos_text in [' ซ้าย-ขวา', ' ซ้าย - ขวา', ' L-R', ' L/R', ' ซ้าย', ' ขวา']:
            thai_clean = thai_clean.replace(pos_text, '')
        thai_clean = thai_clean.strip()

        if not thai_clean:
            continue

        # Parse prices
        replace_val = parse_price(row[col_replace]) if col_replace is not None and col_replace < len(row) else None
        s_val = parse_price(row[col_s]) if col_s is not None and col_s < len(row) else None
        m_val = parse_price(row[col_m]) if col_m is not None and col_m < len(row) else None
        l_val = parse_price(row[col_l]) if col_l is not None and col_l < len(row) else None

        rows_out.append({
            'submodel': submodel,
            'subject': thai_clean,
            'position': position,
            'replace': replace_val,
            's': s_val,
            'm': m_val,
            'l': l_val,
        })

    return rows_out


# ============================================================
# Process each PDF
# ============================================================

PDF_FILES = [
    {
        'path': FORD_DIR + 'All New Ford Ranger.pdf',
        'model': 'All New Ford Ranger',
        'is_suv': False,
        # Pages with price tables (1-indexed): map page -> submodel override
        'page_submodels': {
            2: 'Full',          # Front - Truck (whole body front)
            7: 'Single Cab',    # Center - Single Cab
            10: 'Open Cab',     # Center - Open Cab
            13: 'Double Cab',   # Center - Double Cab
            16: 'Full',         # Rear - Truck (whole body rear)
            20: None,           # Labor list (skip - ราคาค่าแรง misc)
            21: None,           # Labor list cont'd
        }
    },
    {
        'path': FORD_DIR + 'Ford Everest.pdf',
        'model': 'Ford Everest',
        'is_suv': True,
        'page_submodels': {
            2: 'SUV',           # Front - SUV
            5: 'SUV',           # Center - SUV
            8: 'SUV',           # Rear - SUV
            12: None,           # Labor list
            13: None,           # Labor list cont'd
        }
    },
    {
        'path': FORD_DIR + 'Ford Ranger.pdf',
        'model': 'Ford Ranger',
        'is_suv': False,
        'page_submodels': {
            4: 'Full',          # Front - Truck
            10: 'Single Cab',   # Center - Single Cab
            13: 'Open Cab',     # Center - Open Cab
            16: 'Double Cab',   # Center - Double Cab
            19: 'Full',         # Rear - Truck
            23: None,           # Labor list
            24: None,           # Labor list cont'd
        }
    },
    {
        'path': FORD_DIR + 'Next Gen Ford Everest.pdf',
        'model': 'Next Gen Ford Everest',
        'is_suv': True,
        'page_submodels': {
            3: 'SUV',           # Front - Next Gen Ford Everest
            6: 'SUV',           # Center - Next Gen Ford Everest
            10: 'SUV',          # Rear - Next Gen Ford Everest
            14: None,           # Labor list
            15: None,           # Labor list cont'd
        }
    },
]

all_data_rows = []

for pdf_info in PDF_FILES:
    print(f"\nProcessing: {pdf_info['path'].split('/')[-1]}")
    with pdfplumber.open(pdf_info['path']) as pdf:
        for page_1indexed, submodel_override in pdf_info['page_submodels'].items():
            if submodel_override is None:
                # Skip labor misc list
                continue
            page = pdf.pages[page_1indexed - 1]
            tables = page.extract_tables()
            if not tables:
                print(f"  Page {page_1indexed}: No table found, skipping")
                continue
            # Use largest table
            main_table = max(tables, key=lambda t: len(t))
            rows = extract_rows_from_table(main_table, submodel_override, pdf_info['model'])
            print(f"  Page {page_1indexed} [{submodel_override}]: {len(rows)} rows")
            for r in rows:
                r['model'] = pdf_info['model']
            all_data_rows.extend(rows)

print(f"\nTotal rows extracted: {len(all_data_rows)}")

# ============================================================
# Write to Excel - Sheet FORD
# ============================================================

wb = openpyxl.load_workbook(EXCEL_PATH)

# Delete and recreate FORD sheet
if 'FORD' in wb.sheetnames:
    del wb['FORD']

ws_audi = wb['Audi']
ws_ford = wb.create_sheet('FORD')

# Copy column widths from Audi
for col_letter, col_dim in ws_audi.column_dimensions.items():
    ws_ford.column_dimensions[col_letter].width = col_dim.width

# Copy header row from Audi (row 1)
for col in range(1, 13):
    src = ws_audi.cell(1, col)
    dst = ws_ford.cell(1, col)
    dst.value = src.value
    if src.font:
        dst.font = Font(
            name=src.font.name, bold=src.font.bold, size=src.font.size,
            color=src.font.color
        )
    if src.fill and src.fill.fill_type != 'none':
        dst.fill = PatternFill(
            fill_type=src.fill.fill_type,
            fgColor=src.fill.fgColor,
            bgColor=src.fill.bgColor
        )
    if src.alignment:
        dst.alignment = Alignment(
            horizontal=src.alignment.horizontal,
            vertical=src.alignment.vertical,
            wrap_text=src.alignment.wrap_text
        )
    if src.border:
        def copy_side(side):
            if side:
                return Side(style=side.border_style, color=side.color)
            return Side()
        dst.border = Border(
            left=copy_side(src.border.left),
            right=copy_side(src.border.right),
            top=copy_side(src.border.top),
            bottom=copy_side(src.border.bottom)
        )

# Freeze header row
ws_ford.freeze_panes = 'A2'

# Write data rows
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
data_font = Font(name='TH Sarabun New', size=12)
data_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
num_align = Alignment(horizontal='right', vertical='center')

for idx, row_data in enumerate(all_data_rows, start=2):
    row_num = idx - 1  # ลำดับ

    cells = [
        row_num,                    # A: ลำดับ
        'FORD',                     # B: ยี่ห้อ
        row_data['model'],          # C: รุ่น
        row_data['submodel'],       # D: รุ่นย่อย
        row_data['subject'],        # E: รายการ
        row_data['position'],       # F: ตำแหน่ง L/R
        '',                         # G: จังหวัดอู่ซ่อม
        row_data['replace'],        # H: เปลี่ยน
        row_data['s'],              # I: ซ่อมเบา S
        row_data['m'],              # J: ซ่อมกลาง M
        row_data['l'],              # K: ซ่อมหนัก L
        '',                         # L: หมายเหตุ
    ]

    for col, val in enumerate(cells, start=1):
        cell = ws_ford.cell(row=idx, column=col)
        cell.value = val
        cell.font = data_font
        cell.border = border
        if col in (8, 9, 10, 11) and val is not None:
            cell.alignment = num_align
            cell.number_format = '#,##0'
        else:
            cell.alignment = data_align

ws_ford.row_dimensions[1].height = 30

print(f"\nWriting {len(all_data_rows)} rows to Sheet FORD...")
wb.save(EXCEL_PATH)
print("Done! File saved.")
print(f"\nSample rows:")
for r in all_data_rows[:5]:
    print(f"  {r}")
