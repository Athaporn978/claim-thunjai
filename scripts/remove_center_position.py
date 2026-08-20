import os
import openpyxl

def remove_center_from_all_sheets():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)

    total_cleared = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row_idx, column=5).value or '').strip()
            if cell_val == "Center":
                ws.cell(row=row_idx, column=5).value = ""
                total_cleared += 1

    wb.save(excel_path)
    print(f"SUCCESS: Removed 'Center' from Column E across all sheets ({total_cleared} cells cleared) at {excel_path}")

if __name__ == "__main__":
    remove_center_from_all_sheets()
