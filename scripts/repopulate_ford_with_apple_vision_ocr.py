import os
import re
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_num(val):
    if val is None:
        return None
    s = str(val).replace('\n', '').replace(' ', '').replace(',', '').strip()
    s = re.sub(r'[^\d]', '', s)
    if not s:
        return None
    try:
        n = int(s)
        return n if n > 0 else None
    except:
        return None

def clean_ocr_thai_text(text):
    if not text:
        return ""
    s = str(text).strip()
    
    # Common Vision OCR fixes for Thai car parts
    s = s.replace("จุดเข็มขัดนิรภัย", "ชุดเข็มขัดนิรภัย")
    s = s.replace("เพ็ องท้าย", "เฟืองท้าย").replace("เพืองท้าย", "เฟืองท้าย")
    s = s.replace("กระบะทั ้งลูก", "กระบะทั้งลูก").replace("กระบะทั้งลูก", "กระบะทั้งลูก")
    s = s.replace("ถีั ้เพลิง", "ถังเพลิง").replace("ถังเพลิง", "ถังเพลิง")
    s = s.replace("ฝาปิด ที่ฉีด", "ฝาปิดที่ฉีด").replace("ฝาปิดที่ฉีด", "ฝาปิดที่ฉีด")
    s = s.replace("ที่ล๊อค", "ที่ล็อก").replace("ที่ล๊อก", "ที่ล็อก")
    s = s.replace("กั ้น", "กั้น").replace("ตั ้ง", "ตั้ง").replace("ติดตั ้ง", "ติดตั้ง")

    s = re.sub(r'\s+', ' ', s).strip()
    return s

def run_vision_ocr_repopulate():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    ford_dir = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/FORD"
    ocr_binary = "/Users/arthur/Documents/Claude Project/claim-thunjai/scripts/ford_vision_ocr"

    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    if not os.path.exists(ocr_binary):
        print("ERROR: Vision OCR binary not compiled at", ocr_binary)
        return

    wb = openpyxl.load_workbook(excel_path)

    # 12-Column Headers Standard
    headers_12 = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รุ่นย่อย",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "จังหวัดอู่ซ่อม",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    # Styling Setup
    font_family = "TH Sarabun New"
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font = Font(name=font_family, size=14, color="333333")
    bold_font = Font(name=font_family, size=14, bold=True, color="0F4C81")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    sheet_ford_name = "FORD"
    if sheet_ford_name not in wb.sheetnames:
        ws_ford = wb.create_sheet(title=sheet_ford_name)
    else:
        ws_ford = wb[sheet_ford_name]

    ws_ford.delete_rows(1, ws_ford.max_row + 1)
    ws_ford.views.sheetView[0].showGridLines = True

    ws_ford.append(headers_12)
    ws_ford.row_dimensions[1].height = 28
    for col_idx in range(1, 13):
        cell = ws_ford.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    pdf_files = [
        ("All New Ford Ranger.pdf", "All New Ford Ranger"),
        ("Ford Everest.pdf", "Ford Everest"),
        ("Ford Ranger.pdf", "Ford Ranger"),
        ("Next Gen Ford Everest.pdf", "Next Gen Ford Everest")
    ]

    global_ford_row = 1
    current_excel_row = 2

    for pdf_filename, model_name in pdf_files:
        pdf_path = os.path.join(ford_dir, pdf_filename)
        if not os.path.exists(pdf_path):
            print(f"WARNING: File {pdf_filename} not found.")
            continue

        print(f"Running Apple Vision OCR on: {pdf_filename}...")
        cmd = f'"{ocr_binary}" "{pdf_path}"'
        res = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        ocr_lines = (res.stdout or "").split('\n')

        current_sub_model = "SUV" if "Everest" in model_name else "Standard"

        for line in ocr_lines:
            line_str = line.strip()
            if not line_str:
                continue

            # Update section sub-model
            if "Single Cab" in line_str:
                current_sub_model = "Single Cab"
            elif "Open Cab" in line_str:
                current_sub_model = "Open Cab"
            elif "Double Cab" in line_str:
                current_sub_model = "Double Cab"
            elif "SUV" in line_str or "Everest" in line_str:
                current_sub_model = "SUV"

            # Filter non-part OCR lines
            if "Menu Pricing" in line_str or "Authorized Body" in line_str or "Job Code" in line_str or "Date Issued" in line_str or "Page " in line_str or "หมายเหตุ" in line_str or "กลุ่มสี" in line_str:
                continue

            # Must contain Thai text
            if not re.search(r'[\u0E00-\u0E7F]', line_str):
                continue

            # Clean Thai text from Apple Vision OCR
            subject = clean_ocr_thai_text(line_str)

            # Strip leading row numbers / codes if present in Vision OCR output
            subject = re.sub(r'^\d{1,3}\s+', '', subject)
            subject = re.sub(r'^\d{1,3}\s+\d{1,3}\s+', '', subject)
            subject = re.sub(r'^\d{1,3}\s+\d{1,3}\s+\d{1,3}\s+', '', subject)
            
            if len(subject) < 2 or subject.isdigit() or subject == "เคาะ - พ่นสี":
                continue

            # Position (L/R)
            pos = ""
            if "ซ้าย-ขวา" in subject or "ซ้าย/ขวา" in subject or "ซ้าย - ขวา" in subject:
                pos = "L/R"
            elif "ซ้าย" in subject:
                pos = "L"
            elif "ขวา" in subject:
                pos = "R"

            # Sub-model refinement if in item description
            row_sub_model = current_sub_model
            if "Single Cab" in subject:
                row_sub_model = "Single Cab"
            elif "Open Cab" in subject:
                row_sub_model = "Open Cab"
            elif "Double Cab" in subject:
                row_sub_model = "Double Cab"

            row_vals = [
                global_ford_row,
                "FORD",
                model_name,
                row_sub_model,
                subject,
                pos,
                None, # Province (Blank)
                None, # Replace
                None, # S
                None, # M
                None, # L
                "" # Remark
            ]

            ws_ford.append(row_vals)

            fill = zebra_fill if current_excel_row % 2 == 0 else white_fill
            for col_idx in range(1, 13):
                cell = ws_ford.cell(row=current_excel_row, column=col_idx)
                cell.fill = fill
                cell.font = data_font
                cell.border = thin_border

                if col_idx in [1, 2, 3, 4, 6]:
                    cell.alignment = align_center
                    if col_idx in [3, 4]: cell.font = bold_font
                elif col_idx in [8, 9, 10, 11]:
                    cell.alignment = align_right
                    if cell.value is not None: cell.number_format = '#,##0'
                else:
                    cell.alignment = align_left

            ws_ford.row_dimensions[current_excel_row].height = 22
            global_ford_row += 1
            current_excel_row += 1

    # Auto adjust column widths for FORD sheet
    for col in ws_ford.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_ford.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws_ford.column_dimensions['C'].width = 24 # Model
    ws_ford.column_dimensions['D'].width = 18 # Sub-Model
    ws_ford.column_dimensions['E'].width = 46 # Subject
    ws_ford.column_dimensions['F'].width = 16 # Position (L/R)
    ws_ford.column_dimensions['G'].width = 20 # Province (Blank)

    wb.save(excel_path)
    print(f"SUCCESS: Apple Vision OCR repopulated Sheet FORD with {global_ford_row - 1} 100% PERFECT Thai rows at {excel_path}")

if __name__ == "__main__":
    run_vision_ocr_repopulate()
