import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def update_headers_and_split_bmw_mini():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)

    # Styling Setup
    font_family = "TH Sarabun New"
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "จังหวัด",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    # 1. Update Column F header to "จังหวัด" in existing sheets (Audi, BENZ)
    for sheet_name in ["Audi", "BENZ"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.cell(row=1, column=6, value="จังหวัด")
            cell = ws.cell(row=1, column=6)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

    # 2. Remove old combined BMW MINI Cooper sheet if exists
    for s in list(wb.sheetnames):
        if "BMW" in s and "MINI" in s:
            wb.remove(wb[s])

    # 3. Create/Ensure Sheet "BMW" and Sheet "MINI Cooper"
    for target_sheet_name in ["BMW", "MINI Cooper"]:
        if target_sheet_name in wb.sheetnames:
            ws = wb[target_sheet_name]
            ws.delete_rows(1, ws.max_row + 1)
        else:
            ws = wb.create_sheet(title=target_sheet_name)

        ws.views.sheetView[0].showGridLines = True
        ws.append(headers)
        ws.row_dimensions[1].height = 28

        for col_idx in range(1, 12):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

        # Auto adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 6, 16)

    wb.save(excel_path)
    print("SUCCESS: Updated Column F to 'จังหวัด' and created separate BMW and MINI Cooper sheets at", excel_path)

if __name__ == "__main__":
    update_headers_and_split_bmw_mini()
