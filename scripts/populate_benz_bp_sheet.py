import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def populate_benz_clean_bp_only():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)
    
    if "BENZ" not in wb.sheetnames:
        ws = wb.create_sheet(title="BENZ")
    else:
        ws = wb["BENZ"]

    # Clear sheet content to rebuild clean headers and data
    ws.delete_rows(1, ws.max_row + 1)
    ws.views.sheetView[0].showGridLines = True

    # Styling Setup
    font_family = "TH Sarabun New"
    
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font = Font(name=font_family, size=14, color="333333")
    bold_font = Font(name=font_family, size=14, bold=True, color="0F4C81")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    headers = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "คำอธิบาย (Description)",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_idx in range(1, 12):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Single Individual Models Lists
    GA = ["A-Class", "CLA", "GLA"]
    GB = ["C-Class", "E-Class", "GLC"]
    GC = ["S-Class", "GLS"]
    ALL_ABC = ["A-Class", "CLA", "GLA", "C-Class", "E-Class", "GLC", "S-Class", "GLS"]

    # Detailed BP Code Items (9-page Document 100%)
    BENZ_BP_DATA = [
        ("BP0101", "กันชนหน้า", "M", [
            (GA, 10400, 8800, 10400, 12000),
            (GB, 12000, 10400, 12000, 13600),
            (GC, 13600, 12000, 13600, 15200),
        ]),
        ("BP0102", "กาบกันชนหน้า", "M", [
            (GA, 4800, 3520, 4800, 5920),
            (GB, 4800, 4000, 5120, 6400),
            (GC, 6400, 5120, 6400, 7200),
        ]),
        ("BP0103", "กันชนหน้าเสริม (Skirt)", "M", [
            (GA, 5120, 4000, 5120, 5920),
            (GB, 5920, 5120, 5920, 6720),
            (GC, 6720, 5920, 6720, 8000),
        ]),
        ("BP0104", "คิ้วคาดกันชนหน้า L", "L", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2720, 2400, 2720, 3200),
        ]),
        ("BP0105", "คิ้วคาดกันชนหน้า R", "R", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2720, 2400, 2720, 3200),
        ]),
        ("BP0106", "แผ่นรองป้ายทะเบียนหน้า", "M", [(ALL_ABC, 2400, 2400, None, None)]),
        ("BP0107", "ฝาปิดที่ฉีดน้ำไฟใหญ่หน้า", "M", [(ALL_ABC, 800, 800, None, None)]),
        ("BP0108", "ฝาปิดที่ลากรถกันชนหน้า", "M", [(ALL_ABC, 800, 800, None, None)]),
        ("BP0109", "คิ้วตระแกรงช่องลมกันชนหน้า", "M", [(ALL_ABC, 1120, 1120, None, None)]),
        ("BP0201", "บังโคลนหน้า L", "L", [
            (GA, 9600, 8000, 9600, 11200),
            (GB, 11200, 9600, 11200, 12800),
            (GC, 12800, 11200, 12800, 14400),
        ]),
        ("BP0202", "บังโคลนหน้า R", "R", [
            (GA, 9600, 8000, 9600, 11200),
            (GB, 11200, 9600, 11200, 12800),
            (GC, 12800, 11200, 12800, 14400),
        ]),
        ("BP0203", "กาบ+คิ้วคาดบังโคลนหน้า L", "L", [(ALL_ABC, 800, 800, 1120, 1600)]),
        ("BP0204", "กาบ+คิ้วคาดบังโคลนหน้า R", "R", [(ALL_ABC, 800, 800, 1120, 1600)]),
        ("BP0205", "ซับในบังโคลนหน้า L", "L", [
            (GA, 4000, 2720, 4000, 5120),
            (GB, 4320, 3200, 4320, 5600),
            (GC, 5600, 4320, 5600, 6720),
        ]),
        ("BP0206", "ซับในบังโคลนหน้า R", "R", [
            (GA, 4000, 2720, 4000, 5120),
            (GB, 4320, 3200, 4320, 5600),
            (GC, 5600, 4320, 5600, 6720),
        ]),
        ("BP0301", "ฝากระโปรงหน้า", "M", [
            (GA, 11200, 9600, 11200, 12800),
            (GB, 12800, 11200, 12800, 14400),
            (GC, 15200, 13600, 15200, 16800),
        ]),
        ("BP0302", "ขาฝากระโปรงหน้า L", "L", [(ALL_ABC, 1920, 1600, 1920, 2400)]),
        ("BP0303", "ขาฝากระโปรงหน้า R", "R", [(ALL_ABC, 1920, 1600, 1920, 2400)]),
        ("BP0304", "ช่องลมฝากระโปรงหน้า L", "L", [(ALL_ABC, 2720, 2400, 2720, 3200)]),
        ("BP0305", "ช่องลมฝากระโปรงหน้า R", "R", [(ALL_ABC, 2720, 2400, 2720, 3200)]),
        ("BP0306", "ช่องลมฝากระโปรงหน้ากลาง", "M", [(ALL_ABC, 5600, 4800, 5600, 6400)]),
        ("BP0401", "ประตูหน้า L", "L", [
            (GA, None, 8800, 10400, 12000),
            (GB, None, 10400, 12000, 13600),
            (GC, None, 12000, 13600, 15200),
        ]),
        ("BP0401/1", "ประตูหน้า L + รางกระจก (เปลี่ยน)", "L", [
            (GA, 12000, None, None, None),
            (GB, 13600, None, None, None),
            (GC, 15200, None, None, None),
            (["C-Coupe"], 15000, 11000, 13000, 15000),
            (["E-Coupe"], 17000, 13000, 15000, 17000),
        ]),
        ("BP0402", "ประตูหน้า R", "R", [
            (GA, None, 8800, 10400, 12000),
            (GB, None, 10400, 12000, 13600),
            (GC, None, 12000, 13600, 15200),
        ]),
        ("BP0402/1", "ประตูหน้า R + รางกระจก (เปลี่ยน)", "R", [
            (GA, 12000, None, None, None),
            (GB, 13600, None, None, None),
            (GC, 15200, None, None, None),
            (["C-Coupe"], 15000, 11000, 13000, 15000),
            (["E-Coupe"], 17000, 13000, 15000, 17000),
        ]),
        ("BP0403", "มือเปิดประตูหน้า L", "L", [(ALL_ABC, 2400, 1600, 2400, 3200)]),
        ("BP0404", "มือเปิดประตูหน้า R", "R", [(ALL_ABC, 2400, 1600, 2400, 3200)]),
        ("BP0405", "กาบ+คิ้วคาดประตูหน้า L", "L", [
            (GA, 3200, 2720, 3200, 4000),
            (GB, 3520, 3200, 3520, 4320),
            (GC, 4000, 3520, 4000, 4800),
        ]),
        ("BP0406", "กาบ+คิ้วคาดประตูหน้า R", "R", [
            (GA, 3200, 2720, 3200, 4000),
            (GB, 3520, 3200, 3520, 4320),
            (GC, 4000, 3520, 4000, 4800),
        ]),
        ("BP0407", "บานพับประตูหน้า L", "L", [(ALL_ABC, 2400, None, None, None)]),
        ("BP0408", "บานพับประตูหน้า R", "R", [(ALL_ABC, 2400, None, None, None)]),
        ("BP0501", "ประตูหลัง L", "L", [
            (GA, None, 8800, 10400, 12000),
            (GB, None, 10400, 12000, 13600),
            (GC, None, 12000, 13600, 15200),
            (["Vito"], None, 15200, 16800, 20800),
        ]),
        ("BP0501/1", "ประตูหลัง L + รางกระจก (เปลี่ยน)", "L", [
            (GA, 12000, None, None, None),
            (GB, 13600, None, None, None),
            (GC, 15200, None, None, None),
            (["Vito"], 22400, None, None, None),
        ]),
        ("BP0502", "ประตูหลัง R", "R", [
            (GA, None, 8800, 10400, 12000),
            (GB, None, 10400, 12000, 13600),
            (GC, None, 12000, 13600, 15200),
            (["Vito"], None, 15200, 16800, 20800),
        ]),
        ("BP0502/1", "ประตูหลัง R + รางกระจก (เปลี่ยน)", "R", [
            (GA, 12000, None, None, None),
            (GB, 13600, None, None, None),
            (GC, 15200, None, None, None),
            (["Vito"], 22400, None, None, None),
        ]),
        ("BP0503", "มือเปิดประตูหลัง L", "L", [(ALL_ABC, 2400, 1600, 2400, 3200)]),
        ("BP0504", "มือเปิดประตูหลัง R", "R", [(ALL_ABC, 2400, 1600, 2400, 3200)]),
        ("BP0505", "กาบ+คิ้วคาดประตูหลัง L", "L", [
            (GA, 3200, 2720, 3200, 3520),
            (GB, 3520, 3200, 3520, 4000),
            (GC, 4000, 3520, 4000, 4320),
        ]),
        ("BP0506", "กาบ+คิ้วคาดประตูหลัง R", "R", [
            (GA, 3200, 2720, 3200, 3520),
            (GB, 3520, 3200, 3520, 4000),
            (GC, 4000, 3520, 4000, 4320),
        ]),
        ("BP0507", "บานพับประตูหลัง L", "L", [(ALL_ABC, 2400, None, None, None)]),
        ("BP0508", "บานพับประตูหลัง R", "R", [(ALL_ABC, 2400, None, None, None)]),
        ("BP0509", "ซุ้มรับประตูหลัง L", "L", [
            (GA, 4000, 2720, 4000, 5120),
            (GB, 4320, 3200, 4320, 5600),
            (GC, 5600, 4320, 5600, 6720),
        ]),
        ("BP0510", "ซุ้มรับประตูหลัง R", "R", [
            (GA, 4000, 2720, 4000, 5120),
            (GB, 4320, 3200, 4320, 5600),
            (GC, 5600, 4320, 5600, 6720),
        ]),
        ("BP0601", "บังโคลนหลัง L", "L", [
            (GA, 19200, 8800, 10400, 12000),
            (GB, 20800, 10400, 12000, 13600),
            (GC, 22400, 12000, 13600, 15200),
            (["C-Coupe"], 24000, 11000, 13000, 15000),
            (["E-Coupe"], 26000, 13000, 15000, 17000),
        ]),
        ("BP0602", "บังโคลนหลัง R", "R", [
            (GA, 19200, 8800, 10400, 12000),
            (GB, 20800, 10400, 12000, 13600),
            (GC, 22400, 12000, 13600, 15200),
            (["C-Coupe"], 24000, 11000, 13000, 15000),
            (["E-Coupe"], 26000, 13000, 15000, 17000),
        ]),
        ("BP0603", "กาบ+คิ้วคาดบังโคลนหลัง L", "L", [(ALL_ABC, 800, 800, 1120, 1600)]),
        ("BP0604", "กาบ+คิ้วคาดบังโคลนหลัง R", "R", [(ALL_ABC, 800, 800, 1120, 1600)]),
        ("BP0605", "ซับในบังโคลนหลัง L", "L", [
            (GA, 4000, 2720, 4000, 5120),
            (GB, 4320, 3200, 4320, 5600),
            (GC, 5600, 4320, 5600, 6720),
        ]),
        ("BP0606", "ซับในบังโคลนหลัง R", "R", [
            (GA, 4000, 2720, 4000, 5120),
            (GB, 4320, 3200, 4320, 5600),
            (GC, 5600, 4320, 5600, 6720),
        ]),
        ("BP0607", "ฝาปิดถังน้ำมัน", "M", [
            (GA, 1600, 1120, 1600, 1920),
            (GB, 1600, 1120, 1600, 1920),
            (GC, 1600, 1600, 1920, 2400),
            (["Vito"], 2400, 1920, 2400, 2720),
        ]),
        ("BP0701", "ฝากระโปรงหลัง", "M", [
            (GA, 10400, 8800, 10400, 12000),
            (GB, 12000, 10400, 12000, 13600),
            (GC, 13600, 12000, 13600, 15200),
        ]),
        ("BP0702", "ขาฝากระโปรงหลัง L", "L", [(ALL_ABC, 1920, 1600, 1920, 2400)]),
        ("BP0703", "ขาฝากระโปรงหลัง R", "R", [(ALL_ABC, 1920, 1600, 1920, 2400)]),
        ("BP0704", "แผ่นรองป้ายทะเบียนหลัง", "M", [
            (GA, 4800, 4000, 4800, 5600),
            (GB, 5120, 4320, 5120, 5920),
            (GC, 5600, 4800, 5600, 6400),
        ]),
        ("BP0705", "สปอยเลอร์ฝากระโปรงหลัง", "M", [
            (GA, 4800, 4000, 4800, 5600),
            (GB, 5120, 3920, 5120, 5920),
            (GC, 5600, 4800, 5600, 6400),
        ]),
        ("BP0801", "กันชนหลัง", "M", [
            (GA, 9600, 8000, 9600, 11200),
            (GB, 11200, 9600, 11200, 12800),
            (GC, 12800, 11200, 12800, 14400),
        ]),
        ("BP0802", "กาบกันชนหลัง", "M", [
            (GA, 4800, 4000, 4800, 5600),
            (GB, 5120, 4320, 5120, 5920),
            (GC, 5600, 4800, 5600, 6400),
        ]),
        ("BP0803", "คิ้วคาดกันชนหลังกลาง", "M", [
            (GA, 2400, 1920, 2400, 3200),
            (GB, 2400, 1920, 2400, 3200),
            (GC, 3200, 2400, 3200, 4000),
        ]),
        ("BP0804", "คิ้วคาดกันชนหลัง L", "L", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2720, 2400, 2720, 3200),
        ]),
        ("BP0805", "คิ้วคาดกันชนหลัง R", "R", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2720, 2400, 2720, 3200),
        ]),
        ("BP0806", "กันชนหลังเสริม (Skirt)", "M", [
            (GA, 4000, 2720, 4000, 5120),
            (GB, 4000, 2720, 4000, 5120),
            (GC, 4320, 3200, 4320, 5600),
        ]),
        ("BP0807", "ฝาปิดที่ลากรถกันชนหลัง", "M", [(ALL_ABC, 800, 800, None, None)]),
        ("BP0901", "หลังคา", "M", [
            (GA, 20000, 11200, 12800, 14400),
            (GB, 22400, 12800, 14400, 16000),
            (GC, 24800, 14400, 16000, 17600),
            (["Vito"], 48000, 20800, 22400, 24000),
        ]),
        ("BP0902", "กาบหลังคา L", "L", [
            (GA, None, 2720, 3200, 4000),
            (GB, None, 3200, 4000, 4800),
            (GC, None, 4000, 4800, 5600),
        ]),
        ("BP0903", "กาบหลังคา R", "R", [
            (GA, None, 2720, 3200, 4000),
            (GB, None, 3200, 4000, 4800),
            (GC, None, 4000, 4800, 5600),
        ]),
        ("BP0904", "ซันรูฟหลังคา", "M", [(ALL_ABC, 5600, 4800, 5600, 6400)]),
        ("BP0905", "คิ้วหลังคา L", "L", [(ALL_ABC, 2400, 2400, 2400, 2400)]),
        ("BP0906", "คิ้วหลังคา R", "R", [(ALL_ABC, 2400, 2400, 2400, 2400)]),
        ("BP1001", "กาบบันได L", "L", [
            (GA, 5120, 4800, 5120, 5920),
            (GB, 5920, 5120, 5920, 6720),
            (GC, 6720, 5920, 6720, 7520),
        ]),
        ("BP1002", "กาบบันได R", "R", [
            (GA, 5120, 4800, 5120, 5920),
            (GB, 5920, 4646, 5920, 6720),
            (GC, 6720, 5920, 6720, 7520),
        ]),
        ("BP1003", "ฝาปิดรูขึ้นขึ้นแรง", "M", [(ALL_ABC, 800, 800, None, None)]),
        ("BP1004", "เหล็กบันได L", "L", [
            (GA, 5120, 4000, 5120, 6400),
            (GB, 6400, 5120, 6400, 7520),
            (GC, 7520, 6400, 7520, 8320),
        ]),
        ("BP1005", "เหล็กบันได R", "R", [
            (GA, 5120, 4000, 5120, 6400),
            (GB, 6400, 5120, 6400, 7520),
            (GC, 7520, 6400, 7520, 8320),
        ]),
        ("BP1006", "ซับในบันได L", "L", [
            (GA, 3200, 2400, 3200, 4000),
            (GB, 4000, 3200, 4000, 4800),
            (GC, 4800, 4000, 4800, 5600),
        ]),
        ("BP1007", "ซับในบันได R", "R", [
            (GA, 3200, 2400, 3200, 4000),
            (GB, 4000, 3200, 4000, 4800),
            (GC, 4800, 4000, 4800, 5600),
        ]),
        ("BP1101", "ฝาครอบกระจกมองข้าง L", "L", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2400, 1920, 2400, 3200),
        ]),
        ("BP1102", "ฝาครอบกระจกมองข้าง R", "R", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2400, 1920, 2400, 3200),
        ]),
        ("BP1201", "หน้ากระจัง", "M", [(ALL_ABC, 4000, 4000, None, None)]),
        ("BP1301", "คิ้วใต้ไฟใหญ่หน้า L", "L", [(ALL_ABC, 1120, 1120, None, None)]),
        ("BP1302", "คิ้วใต้ไฟใหญ่หน้า R", "R", [(ALL_ABC, 1120, 1120, None, None)]),
        ("BP1401", "แผงรับไฟใหญ่หน้า L (S-Class)", "L", [(["S-Class", "GLS"], 4800, 2400, 3200, 4000)]),
        ("BP1402", "แผงรับไฟใหญ่หน้า R (S-Class)", "R", [(["S-Class", "GLS"], 4800, 2400, 3200, 4000)]),
        ("BP1403", "คานรับฝากระโปรงหน้า", "M", [
            (GA, 2400, 1920, 2400, 3200),
            (GB, 2400, 1920, 2400, 3200),
            (GC, 3200, 2400, 3200, 4000),
        ]),
        ("BP1404", "คานใต้หม้อน้ำ", "M", [(ALL_ABC, 6400, 4800, 5600, 6400)]),
        ("BP1501", "คัสซีหน้า L", "L", [(ALL_ABC, 19200, 7200, 10400, 13600)]),
        ("BP1502", "คัสซีหน้า R", "R", [(ALL_ABC, 19200, 7200, 10400, 13600)]),
        ("BP1503", "คานปิดหัวคัสซีหน้า", "M", [(ALL_ABC, 12800, 6400, 9600, 12800)]),
        ("BP1504", "เหล็กบังฝุ่นหน้า L", "L", [(ALL_ABC, None, 4000, 4800, 5600)]),
        ("BP1505", "เหล็กบังฝุ่นหน้า R", "R", [(ALL_ABC, None, 4000, 4800, 5600)]),
        ("BP1506", "เบ้ายึดโช้คหน้า L", "L", [(ALL_ABC, 11200, 3200, 4800, 8000)]),
        ("BP1507", "เบ้ายึดโช้คหน้า R", "R", [(ALL_ABC, 11200, 3200, 4800, 8000)]),
        ("BP1508", "แผงหลังเครื่อง", "M", [(ALL_ABC, 13600, 5600, 7200, 8000)]),
        ("BP1601", "เสากระจกบังลมหน้า L", "L", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2720, 2400, 2720, 3200),
        ]),
        ("BP1602", "เสากระจกบังลมหน้า R", "R", [
            (GA, 1920, 1600, 1920, 2400),
            (GB, 1920, 1600, 1920, 2400),
            (GC, 2720, 2400, 2720, 3200),
        ]),
        ("BP1603", "เสากลางเก๋ง L", "L", [(ALL_ABC, 6400, 4000, 4800, 6400)]),
        ("BP1604", "เสากลางเก๋ง R", "R", [(ALL_ABC, 6400, 4000, 4800, 6400)]),
        ("BP1605", "ซับในเสากลางเก๋ง L", "L", [(ALL_ABC, 6400, 4000, 4800, 6400)]),
        ("BP1606", "ซับในเสากลางเก๋ง R", "R", [(ALL_ABC, 6400, 4000, 4800, 6400)]),
        ("BP1607", "พื้นวางเท้า L", "L", [(ALL_ABC, 16000, 4800, 8800, 12800)]),
        ("BP1608", "พื้นวางเท้า R", "R", [(ALL_ABC, 16000, 4800, 8800, 12800)]),
        ("BP1609", "แผงเหล็กลำโพงหลัง", "M", [(ALL_ABC, 12800, 6400, 8000, 9600)]),
        ("BP1701", "แผงท้าย", "M", [(ALL_ABC, 12800, 6400, 9600, 12800)]),
        ("BP1702", "ซับในแผงท้าย", "M", [(ALL_ABC, 6400, 3200, 4800, 6400)]),
        ("BP1703", "เบ้ายึดไฟท้าย L", "L", [(ALL_ABC, 6400, 3200, 4800, 6400)]),
        ("BP1704", "เบ้ายึดไฟท้าย R", "R", [(ALL_ABC, 6400, 3200, 4800, 6400)]),
        ("BP1705", "พื้นวางยางอะไหล่", "M", [(ALL_ABC, 9600, 6400, 8000, 9600)]),
        ("BP1706", "คัสซีหลัง L", "L", [(ALL_ABC, 19200, 7200, 10400, 13600)]),
        ("BP1707", "คัสซีหลัง R", "R", [(ALL_ABC, 19200, 7200, 10400, 13600)]),
        ("BP1708", "แผงหลังเก๋ง", "M", [(ALL_ABC, 8000, 3200, 4000, 4800)]),
        ("BP1801", "ล้อแมกซ์หน้า L R (ขอบ 17-18)", "L R", [(ALL_ABC, 2500, None, None, None)]),
        ("BP1801", "ล้อแมกซ์หน้า L R (ขอบ 19-22)", "L R", [(ALL_ABC, 3000, None, None, None)]),
        ("BP1802", "ล้อแมกซ์หลัง L R (ขอบ 17-18)", "L R", [(ALL_ABC, 2500, None, None, None)]),
        ("BP1802", "ล้อแมกซ์หลัง L R (ขอบ 19-22)", "L R", [(ALL_ABC, 3000, None, None, None)]),
        ("BP2000", "ยกเครื่องเข้า-ออก", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2001", "ถอดใส่อุปกรณ์ในห้องเครื่อง", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2002", "ถอดหน้าปัดเรือนไมล์", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2003", "ถอดประกอบด้านหลัง", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2004", "ถอด-ประกอบผ้าหลังคา", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2005", "แว๊คน้ำยาแอร์ R 134 A", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2006", "ถอดใส่อะไหล่ช่วงล่าง", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2007", "ถอด-ประกอบพร้อมเบาะ", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2008", "ถอด-ประกอบคอนโทรล", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2009", "ถอด-ประกอบรังผึ้งแอร์", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2010", "ถอด-ประกอบรังผึ้งหม้อน้ำ", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2011", "ถอด-ประกอบระบบไฟฟ้า", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2012", "ถอด-ประกอบเกียร์", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2015", "ตั้งศูนย์ล้อ", "M", [(ALL_ABC, None, None, None, None)]),
        ("BP2013", "ซีดยางย่อยอะไหล่", "M", [(ALL_ABC, None, 2500, None, None)]),
        ("BP2014", "ถอด-ประกอบกระจก-น้ำยา", "M", [(ALL_ABC, None, None, None, None)]),
    ]

    row_count = 1
    current_excel_row = 2

    # Write BP Code Data (100% 9-page Manual Dataset)
    for code, subject, position_lr, price_groups in BENZ_BP_DATA:
        for model_group, replace_p, s_p, m_p, l_p in price_groups:
            for model_name in model_group:
                row_values = [
                    row_count,
                    "BENZ",
                    model_name, # Single individual model name
                    subject,
                    position_lr,
                    None,
                    replace_p,
                    s_p,
                    m_p,
                    l_p,
                    None
                ]
                ws.append(row_values)
                
                fill = zebra_fill if current_excel_row % 2 == 0 else white_fill
                for col_idx in range(1, 12):
                    cell = ws.cell(row=current_excel_row, column=col_idx)
                    cell.fill = fill
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if col_idx in [1, 2, 3, 5]:
                        cell.alignment = align_center
                        if col_idx == 3: cell.font = bold_font
                    elif col_idx in [7, 8, 9, 10]:
                        cell.alignment = align_right
                        if cell.value is not None: cell.number_format = '#,##0'
                    else:
                        cell.alignment = align_left

                ws.row_dimensions[current_excel_row].height = 22
                row_count += 1
                current_excel_row += 1

    # Auto adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws.column_dimensions['C'].width = 22 # Model
    ws.column_dimensions['D'].width = 34 # Subject
    ws.column_dimensions['E'].width = 16 # Position (L/R)
    ws.column_dimensions['F'].width = 20 # Description (Blank)
    ws.column_dimensions['K'].width = 20 # Remark (Blank)

    wb.save(excel_path)
    print(f"SUCCESS: Clean 9-page BP manual BENZ sheet populated with {row_count - 1} single-model rows at {excel_path}")

if __name__ == "__main__":
    populate_benz_clean_bp_only()
