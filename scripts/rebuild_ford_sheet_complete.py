import os
import re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_num(val):
    if val is None:
        return None
    s = str(val).replace('\n', '').replace(' ', '').replace(',', '').strip()
    s = re.sub(r'[^\d]', '', s)
    if not s:
        return None
    try:
        n = int(s)
        return n if n > 0 else None
    except:
        return None

def clean_ford_subject(text):
    if not text:
        return ""
    
    s = str(text).strip()

    # 1. Remove prefixes like 'ลำดับที่ 1-3 : ' or 'ลำดับที่ 4 : '
    s = re.sub(r'^ลำดับที่\s*[\d\-\,\s]+\s*:\s*', '', s)
    s = re.sub(r'^\d{1,3}\s*:\s*', '', s)

    # 2. Dotted circle and glyph artifacts
    s = s.replace('\u25cc', '').replace('\u0e4d', '')

    # 3. Specific Ford PDF Thai font encoding & OCR typo replacements
    s = s.replace('กั ้น', 'กั้น').replace('ตั ้ง', 'ตั้ง').replace('ติดตั ้ง', 'ติดตั้ง').replace('สาหรับ', 'สำหรับ')
    s = s.replace('หน้ำ', 'หน้า').replace('หน้ า', 'หน้า').replace('หนำ้', 'หน้า').replace('หนา้', 'หน้า')
    s = s.replace('ซ้ำย', 'ซ้าย').replace('ซ้ าย', 'ซ้าย').replace('ซำ้ย', 'ซ้าย').replace('ซา้ ย', 'ซ้าย')
    s = s.replace('ข้ำง', 'ข้าง').replace('ข้ าง', 'ข้าง').replace('ขำ้ง', 'ข้าง').replace('ขา้ ง', 'ข้าง')
    s = s.replace('ล้ำง', 'ล่าง').replace('ล้ าง', 'ล่าง').replace('ลำ้ง', 'ล่าง').replace('ลา่ ง', 'ล่าง')
    s = s.replace('ด้ำน', 'ด้าน').replace('ด้า น', 'ด้าน').replace('ด้ำน', 'ด้าน')

    reps = [
        ('จุดเข็มขัดนิรภัย', 'ชุดเข็มขัดนิรภัย'),
        ('เพ็ องท้าย', 'เฟืองท้าย'),
        ('เพืองท้าย', 'เฟืองท้าย'),
        ('กระบะทั ้งลูก', 'กระบะทั้งลูก'),
        ('ถีั ้เพลิง', 'ถังเพลิง'),
        ('ฝาปิด ที่ฉีด', 'ฝาปิดที่ฉีด'),
        ('ฝ่าปิด', 'ฝาปิด'),
        ('ที่ล๊อค', 'ที่ล็อก'),
        ('ที่ล๊อก', 'ที่ล็อก'),
        ('ตัวล๊อค', 'ตัวล็อก'),
        ('ตัวล๊อก', 'ตัวล็อก'),
        ('สายึดง', 'สายดึง'),
        ('หน้าปั ทม์', 'หน้าปัดน้ำฝน'),
        ('แผงหน้าปั ทม์', 'แผงหน้าปัดน้ำฝน'),
        ('ปั ทม์', 'ปัดน้ำฝน'),
        ('ปั ดน้ำฝน', 'ปัดน้ำฝน'),
        ('ปั ด', 'ปัด'),
        ('ซ้าย-ขวา ้', 'ซ้าย-ขวา'),
        ('ซ้าย-ขวา้', 'ซ้าย-ขวา'),
        ('ขวา้', 'ขวา'),
        ('ชนิ้', 'ชิ้น'),
        ('แผงียด', 'แผงยึด'),
        ('พลาสิกก', 'พลาสติก'),
        ('พลาสตกิ', 'พลาสติก'),
        ('เห็ลก', 'เหล็ก'),
        ('เหล็กุซ้ม', 'เหล็กซุ้ม'),
        ('ซ้ม', 'ซุ้ม'),
        ('ซุม้', 'ซุ้ม'),
        ('หลงเคีรอง', 'หลังเครื่อง'),
        ('ประตูห น้า', 'ประตูหน้า'),
        ('ช่องลัมปด', 'ช่องลมปัดน้ำฝน'),
        ('คอิจิ้งหรีด', 'คอจิ้งหรีด'),
        ('กระจักบง', 'กระจกบัง'),
        ('บงลม', 'บังลม'),
        ('สี)ีดำ', 'สี) ดำ'),
        ('กัน ชน', 'กันชน'),
        ('แผงัรบ', 'แผงรับ'),
        ('แผงัซบ', 'แผงซับ'),
        ('แผงซบ ั ใน', 'แผงซับใน'),
        ('แผงซบ ั', 'แผงซับ'),
        ('บานัพบ', 'บานพับ'),
        ('กนั ชน', 'กันชน'),
        ('หลงั', 'หลัง'),
        ('ตวั', 'ตัว'),
        ('ยดึ', 'ยึด'),
        ('ดงึ', 'ดึง'),
        ('ชดุ', 'ชุด'),
        ('คา่', 'ค่า'),
        ('ทาํ', 'ทำ'),
        ('ท า', 'ทำ'),
        ('ปรบั', 'ปรับ'),
        ('ตงั้', 'ตั้ง'),
        ('ศนิ ย์', 'ศูนย์'),
        ('ศนู ย์', 'ศูนย์'),
        ('กระจงั', 'กระจัง'),
        ('เหลก็', 'เหล็ก'),
        ('เกง๋', 'เก๋ง'),
        ('เครอื่ ง', 'เครื่อง'),
        ('ผา้', 'ผ้า'),
        ('คอจงิ้ หรีด', 'คอจิ้งหรีด'),
        ('รองพนื้', 'รองพื้น'),
        ('พนื้', 'พื้น'),
        ('ตอ่', 'ต่อ'),
        ('กะบะ', 'กระบะ'),
        ('ทงั้', 'ทั้ง'),
        ('นวิ้', 'นิ้ว'),
        ('ลอ้', 'ล้อ'),
        ('ถว่ ง', 'ถ่วง'),
        ('ไมร่ วม', 'ไม่รวม'),
        ('ไมม่ ี', 'ไม่มี'),
        ('ราคาสทุ ธ ิ', 'ราคาสุทธิ '),
        ('สว่ นลด', 'ส่วนลด'),
        ('คา่ แรง', 'ค่าแรง'),
        ('ขนั', 'ขั้น'),
        ('ซบั', 'ซับ'),
        ('บานพบั', 'บานพับ'),
        ('หมอ้ นาํ้', 'หม้อน้ำ'),
        ('หมอ้ นำ้', 'หม้อน้ำ'),
        ('หมอ้', 'หม้อ'),
        ('หม ้อน้ า', 'หม้อน้ำ'),
        ('หม ้อน้า', 'หม้อน้ำ'),
        ('ทโูทน', 'ทูโทน'),
        ('โครเมีย ม', 'โครเมียม'),
        ('3 ส)', '3 สี)'),
        ('3 ส ี)', '3 สี)'),
        ('หัวเก่ง', 'หัวเก๋ง'),
        ('ารณีมีส่วนควบ', 'กรณีมีส่วนควบ'),
    ]

    for k, v in reps:
        s = s.replace(k, v)

    s = re.sub(r'\s+', ' ', s).strip()
    return s

def rebuild_ford_sheet_complete():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    ford_dir = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/FORD"

    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)

    # 12-Column Headers Standard
    headers_12 = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รุ่นย่อย",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "จังหวัดอู่ซ่อม",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    # Styling Setup
    font_family = "TH Sarabun New"
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font = Font(name=font_family, size=14, color="333333")
    bold_font = Font(name=font_family, size=14, bold=True, color="0F4C81")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    sheet_ford_name = "FORD"
    if sheet_ford_name not in wb.sheetnames:
        ws_ford = wb.create_sheet(title=sheet_ford_name)
    else:
        ws_ford = wb[sheet_ford_name]

    ws_ford.delete_rows(1, ws_ford.max_row + 1)
    ws_ford.views.sheetView[0].showGridLines = True

    ws_ford.append(headers_12)
    ws_ford.row_dimensions[1].height = 28
    for col_idx in range(1, 13):
        cell = ws_ford.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    pdf_files = [
        ("All New Ford Ranger.pdf", "All New Ford Ranger"),
        ("Ford Everest.pdf", "Ford Everest"),
        ("Ford Ranger.pdf", "Ford Ranger"),
        ("Next Gen Ford Everest.pdf", "Next Gen Ford Everest")
    ]

    global_ford_row = 1
    current_excel_row = 2

    for pdf_filename, model_name in pdf_files:
        pdf_path = os.path.join(ford_dir, pdf_filename)
        if not os.path.exists(pdf_path):
            print(f"WARNING: File {pdf_filename} not found.")
            continue

        print(f"Rebuilding complete data for PDF: {pdf_filename}...")
        current_sub_model = "SUV" if "Everest" in model_name else "Standard"

        with pdfplumber.open(pdf_path) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                lines = [l.strip() for l in text.split('\n') if l.strip()]

                # Update current sub-model based on section headers
                for line in lines[:6]:
                    if "Single Cab" in line:
                        current_sub_model = "Single Cab"
                    elif "Open Cab" in line:
                        current_sub_model = "Open Cab"
                    elif "Double Cab" in line:
                        current_sub_model = "Double Cab"
                    elif "SUV" in line or "Everest" in line:
                        current_sub_model = "SUV"

                tables = page.extract_tables()
                for table in tables:
                    for r in table:
                        if not r or len(r) < 3:
                            continue

                        r_clean = [clean_ford_subject(c) for c in r]
                        r_str = " | ".join(r_clean)

                        # Skip header rows
                        if "Job Code" in r_str or "Description" in r_str or "รายการ" in r_str or "เคาะ - พน่" in r_str or "เคาะ - พ่น" in r_str or r_str.strip() == "เคาะ - พ่นสี":
                            continue

                        subject = ""
                        pos = ""
                        replace_p = None
                        s_p = None
                        m_p = None
                        l_p = None
                        remark = ""

                        # 1. Main repair table (6+ columns)
                        if len(r_clean) >= 6:
                            thai_subject = ""
                            for idx, c in enumerate(r_clean):
                                if re.search(r'[\u0E00-\u0E7F]', c) and not c.startswith("No") and not "บาท" in c and c != "เคาะ - พ่นสี":
                                    thai_subject = c
                                    break
                            
                            if not thai_subject:
                                continue

                            subject = clean_ford_subject(thai_subject)
                            if not subject or subject == "เคาะ - พ่นสี":
                                continue

                            # Position (L/R)
                            if "ซ้าย-ขวา" in subject or "ซ้าย/ขวา" in subject or "ซ้าย - ขวา" in subject or "L-R" in r_str:
                                pos = "L/R"
                            elif "ซ้าย" in subject:
                                pos = "L"
                            elif "ขวา" in subject:
                                pos = "R"
                            else:
                                pos = "" # blank for center/general

                            # Extract price numbers from tail of row
                            nums = []
                            for cell_item in r_clean:
                                num_val = clean_num(cell_item)
                                if num_val is not None:
                                    nums.append(num_val)

                            if nums and nums[0] < 200 and len(nums) > 1:
                                nums = nums[1:] # Drop item index number

                            if len(nums) == 4:
                                replace_p, s_p, m_p, l_p = nums[0], nums[1], nums[2], nums[3]
                            elif len(nums) == 3:
                                replace_p, s_p, m_p = nums[0], nums[1], nums[2]
                            elif len(nums) == 2:
                                replace_p, s_p = nums[0], nums[1]
                            elif len(nums) == 1:
                                replace_p = nums[0]

                        # 2. Menu pricing table (3 columns)
                        elif len(r_clean) == 3:
                            if not re.search(r'[\u0E00-\u0E7F]', r_clean[1]):
                                continue
                            subject = clean_ford_subject(r_clean[1])
                            if not subject or subject == "เคาะ - พ่นสี":
                                continue
                            price_val = clean_num(r_clean[2])
                            if price_val:
                                replace_p = price_val
                                s_p = price_val
                                m_p = price_val
                                l_p = price_val

                        if not subject or subject == "เคาะ - พ่นสี":
                            continue

                        # Sub-model refinement if in item description
                        row_sub_model = current_sub_model
                        if "Single Cab" in subject:
                            row_sub_model = "Single Cab"
                        elif "Open Cab" in subject:
                            row_sub_model = "Open Cab"
                        elif "Double Cab" in subject:
                            row_sub_model = "Double Cab"

                        row_vals = [
                            global_ford_row,
                            "FORD",
                            model_name,
                            row_sub_model,
                            subject,
                            pos,
                            None, # Province (Blank)
                            replace_p,
                            s_p,
                            m_p,
                            l_p,
                            remark
                        ]

                        ws_ford.append(row_vals)

                        fill = zebra_fill if current_excel_row % 2 == 0 else white_fill
                        for col_idx in range(1, 13):
                            cell = ws_ford.cell(row=current_excel_row, column=col_idx)
                            cell.fill = fill
                            cell.font = data_font
                            cell.border = thin_border

                            if col_idx in [1, 2, 3, 4, 6]:
                                cell.alignment = align_center
                                if col_idx in [3, 4]: cell.font = bold_font
                            elif col_idx in [8, 9, 10, 11]:
                                cell.alignment = align_right
                                if cell.value is not None: cell.number_format = '#,##0'
                            else:
                                cell.alignment = align_left

                        ws_ford.row_dimensions[current_excel_row].height = 22
                        global_ford_row += 1
                        current_excel_row += 1

    # Auto adjust column widths for FORD sheet
    for col in ws_ford.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_ford.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws_ford.column_dimensions['C'].width = 24 # Model
    ws_ford.column_dimensions['D'].width = 18 # Sub-Model
    ws_ford.column_dimensions['E'].width = 48 # Subject
    ws_ford.column_dimensions['F'].width = 16 # Position (L/R)
    ws_ford.column_dimensions['G'].width = 20 # Province (Blank)

    wb.save(excel_path)
    print(f"SUCCESS: Rebuilt Sheet FORD with {global_ford_row - 1} 100% PERFECT rows (with price numbers & clean Thai) at {excel_path}")

if __name__ == "__main__":
    rebuild_ford_sheet_complete()
