import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def add_audi_horizontal_header():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)
    
    if "Audi" not in wb.sheetnames:
        ws = wb.create_sheet(title="Audi")
    else:
        ws = wb["Audi"]

    # Clear sheet content to put clean horizontal header
    ws.delete_rows(1, ws.max_row + 1)
    ws.views.sheetView[0].showGridLines = True

    # Styling
    font_family = "TH Sarabun New"
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Vibrant Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รายการ (Subject)",
        "คำอธิบาย (Description)",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_idx in range(1, 11):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 6, 16)

    wb.save(excel_path)
    print("SUCCESS: Audi horizontal header added successfully at", excel_path)

if __name__ == "__main__":
    add_audi_horizontal_header()
