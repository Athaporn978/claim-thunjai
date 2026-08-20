import os
import openpyxl

def create_blank_sheets_excel():
    target_dir = "/Users/arthur/Documents/Customer/H Tech/Labor Motor"
    excel_path = os.path.join(target_dir, "Labor_Motor_Car_Brands.xlsx")

    # Read subdirectories
    subdirs = [d for d in os.listdir(target_dir) if os.path.isdir(os.path.join(target_dir, d)) and not d.startswith(".")]
    subdirs.sort()

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active

    for idx, brand_folder in enumerate(subdirs):
        # Excel sheet name limit is 31 characters & sanitize invalid chars
        sheet_title = brand_folder.replace(":", " ").replace("/", " ").replace("\\", " ").replace("?", "").replace("*", "").replace("[", "").replace("]", "")
        sheet_title = " ".join(sheet_title.split())[:31]

        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        # Sheet is left completely blank inside (0 text/data added)

    # Remove default Sheet if present
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(excel_path)
    print("SUCCESS: Blank sheets Excel saved at", excel_path)

if __name__ == "__main__":
    create_blank_sheets_excel()
