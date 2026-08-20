import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def populate_benz_data():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)
    
    if "BENZ" not in wb.sheetnames:
        ws = wb.create_sheet(title="BENZ")
    else:
        ws = wb["BENZ"]

    # Clear sheet content to rebuild clean headers and data
    ws.delete_rows(1, ws.max_row + 1)
    ws.views.sheetView[0].showGridLines = True

    # Styling Setup
    font_family = "TH Sarabun New"
    
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font = Font(name=font_family, size=14, color="333333")
    bold_font = Font(name=font_family, size=14, bold=True, color="0F4C81")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Headers 100% identical to Audi sheet
    headers = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "คำอธิบาย (Description)",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_idx in range(1, 12):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Benz Items Data from Talingchan document (TLC. 06/2566)
    # Format: (Subject, Position, [ (Model, Replace, S, M, L), ... ])
    BENZ_RAW_DATA = [
        ("กันชนหน้า", "Center", [
            ("C-Class", 12750, 11250, 12750, 14250),
            ("E-Class", 14250, 12750, 14250, 15000),
            ("S-Class", 15750, 14250, 15750, 18750),
        ]),
        ("บังโคลนหน้า", "L/R", [
            ("C-Class", 12000, 10500, 12000, 13500),
            ("E-Class", 13500, 12000, 13500, 15000),
            ("S-Class", 15000, 13500, 15000, 18000),
        ]),
        ("ฝากระโปรงหน้า", "Center", [
            ("C-Class", 13500, 12000, 13500, 15000),
            ("E-Class", 15000, 13500, 15000, 16500),
            ("S-Class", 18000, 16500, 18000, 19500),
        ]),
        ("ประตูหน้า", "L/R", [
            ("C-Class", 12750, 11250, 12750, 14250),
            ("E-Class", 14250, 12750, 14250, 15750),
            ("S-Class", 15750, 14250, 15750, 18000),
        ]),
        ("บังโคลนหลัง", "L/R", [
            ("C-Class", 24000, 11250, 12750, 14250),
            ("E-Class", 25500, 12750, 14250, 16500),
            ("S-Class", 27000, 14250, 15750, 18750),
        ]),
        ("ฝากระโปรงหลัง", "Center", [
            ("C-Class", 12750, 11250, 12750, 14250),
            ("E-Class", 14250, 12750, 14250, 15750),
            ("S-Class", 15750, 14250, 15750, 15750),
        ]),
        ("กันชนหลัง", "Center", [
            ("C-Class", 12000, 10500, 12000, 13500),
            ("E-Class", 13500, 12000, 13500, 15000),
            ("S-Class", 15000, 13500, 15000, 18000),
        ]),
        ("ประตูหลัง", "L/R", [
            ("C-Class", 12750, 11250, 12750, 14250),
            ("E-Class", 14250, 12750, 14250, 15750),
            ("S-Class", 17250, 14250, 15750, 18750),
        ]),
        ("หลังคา", "Center", [
            ("C-Class", 24750, 13500, 15000, 16500),
            ("E-Class", 27000, 14250, 15750, 18000),
            ("S-Class", 30000, 16500, 18750, 21000),
        ]),
        ("กาบบันได", "L/R", [
            ("C-Class", 6000, 6000, 6750, 7500),
            ("E-Class", 6000, 6000, 6750, 7500),
            ("S-Class", 6000, 6000, 6750, 7500),
        ]),
    ]

    row_count = 1
    current_excel_row = 2

    for subject, position_lr, model_rows in BENZ_RAW_DATA:
        for model_name, replace_p, s_p, m_p, l_p in model_rows:
            row_values = [
                row_count,
                "BENZ",
                model_name,
                subject,
                position_lr,
                None, # Column F: Description (Blank)
                replace_p,
                s_p,
                m_p,
                l_p,
                None  # Column K: Remark (Blank)
            ]
            ws.append(row_values)
            
            # Format cell
            fill = zebra_fill if current_excel_row % 2 == 0 else white_fill
            for col_idx in range(1, 12):
                cell = ws.cell(row=current_excel_row, column=col_idx)
                cell.fill = fill
                cell.font = data_font
                cell.border = thin_border
                
                if col_idx in [1, 2, 3, 5]: # No, Brand, Model, Position
                    cell.alignment = align_center
                    if col_idx == 3:
                        cell.font = bold_font
                elif col_idx in [7, 8, 9, 10]: # Prices (Replace, S, M, L)
                    cell.alignment = align_right
                    if cell.value is not None:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = align_left

            ws.row_dimensions[current_excel_row].height = 22
            row_count += 1
            current_excel_row += 1

    # Auto adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws.column_dimensions['D'].width = 28 # Subject
    ws.column_dimensions['E'].width = 16 # Position (L/R)
    ws.column_dimensions['F'].width = 22 # Description (Blank)
    ws.column_dimensions['K'].width = 22 # Remark (Blank)

    wb.save(excel_path)
    print(f"SUCCESS: Populated BENZ sheet with {row_count - 1} detailed rows at {excel_path}")

if __name__ == "__main__":
    populate_benz_data()
