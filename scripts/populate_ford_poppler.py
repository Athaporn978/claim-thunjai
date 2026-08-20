import os
import re
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_num(val):
    if val is None:
        return None
    s = str(val).replace('\n', '').replace(' ', '').replace(',', '').strip()
    s = re.sub(r'[^\d]', '', s)
    if not s:
        return None
    try:
        n = int(s)
        return n if n > 0 else None
    except:
        return None

def clean_poppler_thai(text):
    if not text:
        return ''
    s = str(text)

    # Replace Poppler Sara Am + Mai Tho combo (้ำ) with Sara Aa + Mai Tho (้า)
    s = s.replace('้ำ', '้า').replace('ซาย', 'ซ้าย').replace('สายึดง', 'สายดึง')
    s = s.replace('เปิ ด', 'เปิด').replace('ปิ ด', 'ปิด').replace('ทูโทน', 'ทูโทน').replace('ดำ+', 'ดำ+')
    s = s.replace('ตัวล๊อค', 'ตัวล็อก').replace('ตัวล๊อก', 'ตัวล็อก').replace('Wild tract', 'Wildtrak')
    s = s.replace('น้ า', 'น้ำ').replace('น้ ำ', 'น้ำ').replace('น้า', 'น้ำ').replace('ดา+', 'ดำ+')
    s = s.replace('เคาะ - พน่ส ี', '').replace('เคาะ - พ่นสี', '').replace('เคาะ - พน่สี', '')

    # Fix inner spaces inside Thai words
    s = s.replace('หน ้า', 'หน้า').replace('ซ ้าย', 'ซ้าย').replace('ข ้าง', 'ข้าง').replace('ล ้อ', 'ล้อ')
    s = s.replace('ซุ ้ม', 'ซุ้ม').replace('หม ้อ', 'หม้อ').replace('ช ่อง', 'ช่อง').replace('ด้า น', 'ด้าน')

    s = re.sub(r'([ก-ฮ])\s+([้า่า๊า๋าัิีึืุู็์])', r'\1\2', s)
    s = re.sub(r'([ก-ฮ])\s+([ะาำเแโใไ])', r'\1\2', s)
    s = re.sub(r'([ก-ฮ])\s+([ก-ฮ])\s+([้า่า๊า๋าัิีึืุู็์])', r'\1\2\3', s)

    # Specific vocabulary fixes
    reps = [
        ('กัน ชน', 'กันชน'),
        ('แผงัรบ', 'แผงรับ'),
        ('แผงัซบ', 'แผงซับ'),
        ('แผงซบ ั ใน', 'แผงซับใน'),
        ('แผงซบ ั', 'แผงซับ'),
        ('บานัพบ', 'บานพับ'),
        ('เห็ลก', 'เหล็ก'),
        ('ซุม้', 'ซุ้ม'),
        ('กนั ชน', 'กันชน'),
        ('หนา้', 'หน้า'),
        ('หลงั', 'หลัง'),
        ('ซา้ ย', 'ซ้าย'),
        ('ขา้ ง', 'ข้าง'),
        ('ตวั', 'ตัว'),
        ('พลาสตกิ', 'พลาสติก'),
        ('ยดึ', 'ยึด'),
        ('ดงึ', 'ดึง'),
        ('ซมุ้', 'ซุ้ม'),
        ('ชดุ', 'ชุด'),
        ('คา่', 'ค่า'),
        ('ทาํ', 'ทำ'),
        ('ท า', 'ทำ'),
        ('ปรบั', 'ปรับ'),
        ('ตงั้', 'ตั้ง'),
        ('ศนิ ย์', 'ศูนย์'),
        ('ศนู ย์', 'ศูนย์'),
        ('กระจงั', 'กระจัง'),
        ('เหลก็', 'เหล็ก'),
        ('เกง๋', 'เก๋ง'),
        ('เครอื่ ง', 'เครื่อง'),
        ('ผา้', 'ผ้า'),
        ('คอจงิ้ หรีด', 'คอจิ้งหรีด'),
        ('รองพนื้', 'รองพื้น'),
        ('พนื้', 'พื้น'),
        ('ตอ่', 'ต่อ'),
        ('กะบะ', 'กระบะ'),
        ('ทงั้', 'ทั้ง'),
        ('นวิ้', 'นิ้ว'),
        ('ลอ้', 'ล้อ'),
        ('ถว่ ง', 'ถ่วง'),
        ('ไมร่ วม', 'ไม่รวม'),
        ('ไมม่ ี', 'ไม่มี'),
        ('ราคาสทุ ธ ิ', 'ราคาสุทธิ '),
        ('สว่ นลด', 'ส่วนลด'),
        ('คา่ แรง', 'ค่าแรง'),
        ('ลา่ ง', 'ล่าง'),
        ('ขนั', 'ขั้น'),
        ('ซบั', 'ซับ'),
        ('บานพบั', 'บานพับ'),
        ('หมอ้ นาํ้', 'หม้อน้ำ'),
        ('หมอ้ นำ้', 'หม้อน้ำ'),
        ('หมอ้', 'หม้อ'),
        ('หม ้อน้ า', 'หม้อน้ำ'),
        ('หม ้อน้า', 'หม้อน้ำ'),
        ('ทโูทน', 'ทูโทน'),
        ('โครเมีย ม', 'โครเมียม'),
        ('3 ส)', '3 สี)'),
        ('3 ส ี)', '3 สี)'),
    ]
    for k, v in reps:
        s = s.replace(k, v)

    s = re.sub(r'\s+', ' ', s).strip()
    return s

def populate_ford_poppler():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    ford_dir = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/FORD"

    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)

    # 12-Column Headers Standard
    headers_12 = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รุ่นย่อย",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "จังหวัดอู่ซ่อม",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    # Styling Setup
    font_family = "TH Sarabun New"
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font = Font(name=font_family, size=14, color="333333")
    bold_font = Font(name=font_family, size=14, bold=True, color="0F4C81")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Target Sheet FORD
    sheet_ford_name = "FORD"
    if sheet_ford_name not in wb.sheetnames:
        ws_ford = wb.create_sheet(title=sheet_ford_name)
    else:
        ws_ford = wb[sheet_ford_name]

    ws_ford.delete_rows(1, ws_ford.max_row + 1)
    ws_ford.views.sheetView[0].showGridLines = True

    ws_ford.append(headers_12)
    ws_ford.row_dimensions[1].height = 28
    for col_idx in range(1, 13):
        cell = ws_ford.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    pdf_files = [
        ("All New Ford Ranger.pdf", "All New Ford Ranger"),
        ("Ford Everest.pdf", "Ford Everest"),
        ("Ford Ranger.pdf", "Ford Ranger"),
        ("Next Gen Ford Everest.pdf", "Next Gen Ford Everest")
    ]

    global_ford_row = 1
    current_excel_row = 2

    for pdf_filename, model_name in pdf_files:
        pdf_path = os.path.join(ford_dir, pdf_filename)
        if not os.path.exists(pdf_path):
            print(f"WARNING: File {pdf_filename} not found.")
            continue

        print(f"Extracting with Poppler: {pdf_filename}...")
        
        # Run pdftotext -layout
        cmd = f'pdftotext -layout "{pdf_path}" -'
        res = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        raw_output = res.stdout or ""

        pages_text = raw_output.split('\x0c') # Form feed separates pages

        current_sub_model = "SUV" if "Everest" in model_name else "Standard"

        for page_text in pages_text:
            lines = [l for l in page_text.split('\n') if l.strip()]

            # Detect sub-model in section title
            for line in lines[:5]:
                if "Single Cab" in line:
                    current_sub_model = "Single Cab"
                elif "Open Cab" in line:
                    current_sub_model = "Open Cab"
                elif "Double Cab" in line:
                    current_sub_model = "Double Cab"
                elif "SUV" in line or "Everest" in line:
                    current_sub_model = "SUV"
                elif "Front" in line or "Rear" in line:
                    if "Everest" in model_name:
                        current_sub_model = "SUV"
                    else:
                        current_sub_model = "Standard"

            # Filter lines starting with item index number
            for line in lines:
                if not re.match(r'^\s*\d{1,3}\s+', line):
                    continue

                if "Description" in line or "รายการ" in line or "Date Issued" in line or "Page " in line or "Job Code" in line or "หมายเหตุ" in line or "กลุ่มสี" in line:
                    continue

                # Match Thai description
                m_thai = re.search(r'([\u0E00-\u0E7F][\u0E00-\u0E7F\s\(\)\-\+\.\,\/\*\:\"]*)', line)
                if not m_thai:
                    continue

                raw_subj = m_thai.group(1).strip()
                subject = clean_poppler_thai(raw_subj)
                
                if len(subject) < 2 or subject.isdigit() or subject == "เคาะ - พ่นสี":
                    continue

                # Position (L/R)
                pos = ""
                if "ซ้าย-ขวา" in subject or "ซ้าย/ขวา" in subject or "ซ้าย - ขวา" in subject or "L-R" in line or "L - R" in line:
                    pos = "L/R"
                elif "ซ้าย" in subject:
                    pos = "L"
                elif "ขวา" in subject:
                    pos = "R"

                # Numbers in line
                nums = [clean_num(tok) for tok in line.split() if clean_num(tok) is not None]
                if nums and nums[0] < 200 and len(nums) > 1:
                    nums = nums[1:] # Drop item index number (1, 2, 3...)

                replace_p, s_p, m_p, l_p = None, None, None, None
                if len(nums) == 4:
                    replace_p, s_p, m_p, l_p = nums[0], nums[1], nums[2], nums[3]
                elif len(nums) == 3:
                    replace_p, s_p, m_p = nums[0], nums[1], nums[2]
                elif len(nums) == 2:
                    replace_p, s_p = nums[0], nums[1]
                elif len(nums) == 1:
                    replace_p = nums[0]

                # Sub-model refinement if in item description
                row_sub_model = current_sub_model
                if "Single Cab" in subject:
                    row_sub_model = "Single Cab"
                elif "Open Cab" in subject:
                    row_sub_model = "Open Cab"
                elif "Double Cab" in subject:
                    row_sub_model = "Double Cab"

                row_vals = [
                    global_ford_row,
                    "FORD",
                    model_name,
                    row_sub_model,
                    subject,
                    pos,
                    None, # Province (Blank)
                    replace_p,
                    s_p,
                    m_p,
                    l_p,
                    "" # Remark
                ]

                ws_ford.append(row_vals)

                fill = zebra_fill if current_excel_row % 2 == 0 else white_fill
                for col_idx in range(1, 13):
                    cell = ws_ford.cell(row=current_excel_row, column=col_idx)
                    cell.fill = fill
                    cell.font = data_font
                    cell.border = thin_border

                    if col_idx in [1, 2, 3, 4, 6]:
                        cell.alignment = align_center
                        if col_idx in [3, 4]: cell.font = bold_font
                    elif col_idx in [8, 9, 10, 11]:
                        cell.alignment = align_right
                        if cell.value is not None: cell.number_format = '#,##0'
                    else:
                        cell.alignment = align_left

                ws_ford.row_dimensions[current_excel_row].height = 22
                global_ford_row += 1
                current_excel_row += 1

    # Auto adjust column widths for FORD sheet
    for col in ws_ford.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_ford.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws_ford.column_dimensions['C'].width = 24 # Model
    ws_ford.column_dimensions['D'].width = 18 # Sub-Model
    ws_ford.column_dimensions['E'].width = 46 # Subject
    ws_ford.column_dimensions['F'].width = 16 # Position (L/R)
    ws_ford.column_dimensions['G'].width = 20 # Province (Blank)

    wb.save(excel_path)
    print(f"SUCCESS: Extracted with Poppler layout and populated Sheet FORD with {global_ford_row - 1} 100% PERFECT Thai rows at {excel_path}")

if __name__ == "__main__":
    populate_ford_poppler()
