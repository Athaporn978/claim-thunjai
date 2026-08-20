import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def populate_bmw_and_mini_all_severity_cols():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)
    
    # Ensure sheets
    for s_name in ["BMW", "MINI Cooper"]:
        if s_name not in wb.sheetnames:
            ws = wb.create_sheet(title=s_name)
        else:
            ws = wb[s_name]

        ws.delete_rows(1, ws.max_row + 1)
        ws.views.sheetView[0].showGridLines = True

    ws_bmw = wb["BMW"]
    ws_mini = wb["MINI Cooper"]

    # Styling Setup
    font_family = "TH Sarabun New"
    
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font = Font(name=font_family, size=14, color="333333")
    bold_font = Font(name=font_family, size=14, bold=True, color="0F4C81")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    headers = [
        "ลำดับ (No.)",
        "ยี่ห้อ (Brand)",
        "รุ่น (Model)",
        "รายการ (Subject)",
        "ตำแหน่ง (L/R)",
        "จังหวัดอู่ซ่อม",
        "เปลี่ยน (Replace)",
        "ซ่อมเบา (S)",
        "ซ่อมกลาง (M)",
        "ซ่อมหนัก (L)",
        "หมายเหตุ (Remark)"
    ]

    for ws in [ws_bmw, ws_mini]:
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, 12):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

    # BMW Models Groups Unpacked
    BMW_G1 = ["Serie 1", "Serie 3", "Coupe"]
    BMW_G2 = ["Serie 5", "Serie 6", "Z-Series", "X3"]
    BMW_G3 = ["Serie 7", "X5", "X6"]

    REMARK_COLOR_PEARL = "สีมุก เพิ่มราคาค่าทำสี 30%"

    # Complete 99 Items from PDF (Item No, Subject, Position, Replace_Price, MINI_P, BMW1_P, BMW2_P, BMW3_P)
    PDF_ITEMS_99 = [
        (1, "กันชนหน้า", "Center", None, 5890, 5890, 7180, 8830),
        (2, "กันชนหลัง", "Center", None, 5890, 5890, 7180, 8830),
        (3, "ฝากระโปรงหน้า", "Center", 1500, 6480, 6480, 7180, 8830),
        (4, "ฝากระโปรงหลัง", "Center", 1500, 5890, 6480, 7180, 8830),
        (5, "บังโคลนหน้า ซ้าย, ขวา", "L/R", None, 2950, 5890, 6480, 8830),
        (6, "บังโคลนหลัง ซ้าย, ขวา", "L/R", None, 5890, 5890, 6480, 8830),
        (7, "ประตูหน้า ซ้าย, ขวา", "L/R", None, 5890, 5890, 6480, 7180),
        (8, "ประตูหลัง ซ้าย, ขวา", "L/R", None, None, 5890, 6480, 7180),
        (9, "แผงท้าย (ทั้งชุด)", "Center", None, 4120, 4600, 5180, 5890),
        (10, "หลังคา", "Center", None, 7770, 7770, 8480, 9070),
        (11, "กาบกันชน หน้า, หลัง (สเกิร์ต คิ้ว)", "Center", None, 2590, 2590, 3070, 3180),
        (12, "โครงในกันชน หน้า, หลัง", "Center", None, None, 2010, 2240, 2590),
        (13, "แผงใต้กันชน หน้า, หลัง", "Center", None, 2360, 2590, 2950, 3180),
        (14, "ขากันชน หน้า, หลัง", "Center", None, 1180, 1300, 1650, 2010),
        (15, "โครงกระจังหน้า", "Center", None, 2950, 3300, 3650, 3890),
        (16, "คิ้วกระจังหน้า", "Center", None, 1770, 2010, 2360, 2590),
        (17, "คิ้วใต้ไฟหน้า ซ้าย, ขวา", "L/R", None, None, 1300, 1300, 2010),
        (18, "คิ้วใต้ไฟหลัง ซ้าย, ขวา", "L/R", None, None, 1650, 1770, 2010),
        (19, "เบ้าไฟหน้า ซ้าย, ขวา", "L/R", None, None, 2010, 2010, 2590),
        (20, "เบ้าไฟหลัง ซ้าย, ขวา", "L/R", None, 1770, 2010, 2010, 2590),
        (21, "แผงรับไฟหน้า ซ้าย, ขวา", "L/R", None, 2010, 2010, 2360, 2590),
        (22, "คานรับหม้อนํ้าแผงหน้า", "Center", None, 2360, 2590, 2950, 3180),
        (23, "คานรับฝากระโปรงหน้า, หลัง", "Center", None, 2360, 2590, 2950, 3180),
        (24, "แผงรับไฟหน้าหรือแผงหน้าทั้งชุด", "Center", None, 3300, 3650, 3890, 4240),
        (25, "แผงหลังเครื่อง", "Center", None, 4120, 4600, 5180, 5890),
        (26, "แชชซีส์ หน้า หลัง ซ้าย-ขวา", "L/R", None, 5890, 6480, 6480, 6480),
        (27, "คานปิดหัวแชชซีส์ หน้า หลัง", "Center", None, 5890, 6480, 6480, 6480),
        (28, "บังฝุ่นเหล็กหน้า ซ้าย ขวา", "L/R", None, 2830, 3180, 3650, 3890),
        (29, "บังฝุ่นเหล็กหลัง ซ้าย ขวา", "L/R", None, 2830, 3180, 3650, 3890),
        (30, "แผงใต้กระจกหน้า (คอจิ้งหรีด)", "Center", None, 2950, 2360, 2590, 2950),
        (31, "เหล็กรับกลอนฝากระโปรงหน้า", "Center", None, 1180, 1300, 1300, 1650),
        (32, "ขาฝากระโปรงหน้า ซ้าย ขวา", "L/R", None, 1180, 1300, 1300, 1650),
        (33, "ขาฝากระโปรงหลัง ซ้าย ขวา", "L/R", None, 1180, 1300, 1300, 1650),
        (34, "คิ้ววงเดือนหน้า หลัง (คิ้วบังโคลน)", "Center", None, 2120, 1650, 1650, 1890),
        (35, "ยางกันโคลน", "Center", None, None, 1650, 1650, 1890),
        (36, "ซับในบังโคลนหน้า ซ้าย ขวา", "L/R", None, 2830, 3180, 3420, 3650),
        (37, "โครงในแผงใต้กระจกบังลมหน้า", "Center", None, 4120, 4600, 5180, 5890),
        (38, "แผงต่อฝากระโปรงหน้า", "Center", None, 4120, 4600, 5180, 5890),
        (39, "เบ้ามือเปิดประตู", "Center", None, 950, 1060, 1060, 1300),
        (40, "ฝาครอบกระจกมองข้าง ขาซ้าย ขวา", "L/R", None, 1300, 1300, 1300, 1650),
        (41, "บานพับประตู", "Center", None, 950, 1060, 1060, 1060),
        (42, "กาบประตู คิ้วคาด", "L/R", None, None, 1890, 2120, 2360),
        (43, "โครงในประตู", "L/R", None, 2830, 3180, 3650, 3890),
        (44, "บันไดด้าน ซ้าย ขวา", "L/R", None, 3650, 3650, 3890, 4240),
        (45, "ซับโครงในบันได ซ้าย ขวา", "L/R", None, 2360, 2360, 2590, 2950),

        # Page 2 Items
        (46, "สเกิร์ตบันได (กาบ) ซ้าย ขวา", "L/R", None, 3650, 3180, 3650, 3890),
        (47, "คันปรับตั้งศูนย์โครงตัวเก๋ง", "Center", None, 8830, 9770, 10360, 11660),
        (48, "พื้นวางเท้าหน้า หลัง ซ้าย ขวา", "L/R", None, 2590, 2950, 3180, 3420),
        (49, "ซุ้มกระโหลกเกียร์", "Center", None, 2830, 3180, 3650, 3890),
        (50, "คานรับพื้นตัวยาว ซ้าย ขวา", "L/R", None, 3540, 3890, 4360, 4600),
        (51, "คานรับพื้นตัวขวาง ซ้าย ขวา", "L/R", None, 2830, 3180, 3650, 3890),
        (52, "โครงยึดหน้าปัด", "Center", None, 2360, 2590, 2950, 3180),
        (53, "แป้นยึดโครงหน้าปัด", "Center", None, 1180, 1300, 1300, 1650),
        (54, "แป้นยึดขาเบรค", "Center", None, 1180, 1300, 1300, 1650),
        (55, "เบ้ายึดโช้ค ซ้าย ขวา", "L/R", None, 2830, 3180, 3420, 3650),
        (56, "ดันตั้งศูนย์ห้องเครื่องโย้ บิด", "Center", None, 8830, 8830, 8830, 9420),
        (57, "ดัด ปรับ ตั้งศูนย์โครงท้าย", "Center", None, None, 5180, 5890, 6480),
        (58, "เบ้าคอถังน้ำมัน", "Center", None, 1650, 1890, 2120, 2360),
        (59, "ฝาปิดถังน้ำมัน", "Center", None, 1060, 1060, 1300, 1300),
        (60, "สปอยร์เลอร์ หน้า หลัง", "Center", None, 3650, 3650, 3650, 3890),
        (61, "พื้นในท้าย ซ้าย ขวา", "L/R", None, 2830, 3180, 3420, 3650),
        (62, "พื้นในช่องวางยางอะไหล่", "Center", None, 2830, 3180, 3420, 3650),
        (63, "ซับในแผงท้าย", "Center", None, 2360, 2590, 2590, 2950),
        (64, "ซับรูฟหลังคา", "Center", None, 3540, 3180, 3420, 3650),
        (65, "ชิ้นข้างหลังคา ซ้าย ขวา", "L/R", None, None, 2950, 3180, 3420),
        (66, "เสากระจกบังลมหน้า ซ้าย ขวา", "L/R", None, 1770, 2010, 2590, 2950),
        (67, "เสากระจกบังลมหลัง ซ้าย ขวา", "L/R", None, 2830, 3180, 3420, 3650),
        (68, "ซ้บในเสาหน้า กลาง หลัง", "Center", None, 2830, 1890, 2120, 2360),
        (69, "เสากลางเก๋ง ซ้าย ขวา", "L/R", None, 2830, 3180, 3420, 3650),
        (70, "เสาประตูหน้า ซ้าย ขวา", "L/R", None, 1650, 1890, 2120, 2360),
        (71, "เสาประตูหลัง ซ้าย ขวา", "L/R", None, None, 1890, 2120, 2360),
        (72, "สปอยร์เลอร์ ฝาท้าย BM ตูดเป็ด", "Center", None, 3300, 3300, 3300, 3300),
        (73, "โช๊คกันชนหน้า ซ้าย ขวา", "L/R", None, 1770, 2010, 2120, 2360),
        (74, "โช๊คกันชนหลัง ซ้าย ขวา", "L/R", None, None, 1890, 2120, 2360),
        (75, "แผ่นรองป้ายทะเบียนหน้า", "Center", None, 1650, 1650, 2240, 2950),
        (76, "แผ่นรองป้ายทะเบียนหลัง", "Center", None, None, 2360, 2590, 2950),
        (77, "ล้อแม็ก มาตรฐาน", "Center", None, 2360, 2590, 3070, 3650),
        (78, "ครีบปลาฉลาม", "Center", None, 950, 1060, 1770, 2360),
        (79, "ฝาครอบไฟสปอร์ตไลท์", "Center", None, None, 1060, 1060, 1060),
        (80, "แผ่นปิดใต้กันชนท้าย", "Center", None, 2710, 3070, 3070, 3070),
        (81, "ชุดแอร์โร MINI", "Center", None, 22370, None, None, None),
        (82, "ล้อแม็ก 17\" 19\" 20\" 22\"", "Center", None, 2950, 3650, 4120, 4710),
        (83, "จมูกฝากระโปรงหน้า", "Center", None, 2120, None, None, None),
        (84, "กระจังหน้า", "Center", None, 2950, None, None, None),
        (85, "ฝาครอบกระจกมองข้าง", "Center", None, 1300, None, None, None),
        (86, "ฝาครอบกระจกมองข้าง+ฐาน ข้างละ", "Center", None, 2360, None, None, None),
        (87, "ติดตั้งกระจกมองข้าง (กรณีของแถม)", "Center", None, 590, None, None, None),
        (88, "คิ้วใต้กระจัง (ติดกันชน) สีตัวรถ", "Center", None, 1180, None, None, None),
        (89, "สปอยร์เลอร์ใต้กันชนหน้า (สีตัวรถ)", "Center", None, 3300, 3300, 3300, 3300),

        # Page 3 Items
        (90, "สปอยร์เลอร์ใต้กันชนหลัง (สีตัวรถ)", "Center", None, 3300, 3300, 3300, 3300),
        (91, "คิ้วล้อ 4 ล้อ ข้างละ 1,800บาท", "Center", None, 8480, None, None, None),
        (92, "ค่าติดตั้งสปอยร์เลอร์ฝาท้าย", "Center", None, 8240, None, None, None),
        (93, "ทำสีตัวถังกาบบันได ซ้าย ขวา ข้างละ", "L/R", None, 3540, None, None, None),
        (94, "แป๊บยึดขาฝากระโปรงหน้า", "Center", None, None, 2590, 2950, 3180),
        (95, "แผงปิดไฟใหญ่หน้า ไฟป๊อป", "Center", None, None, 1890, 2120, 2360),
        (96, "แผงหลังเก๋ง (ใต้กระจกบังลมหลัง)", "Center", None, None, 3180, 3420, 3650),
        (97, "แผงลำโพง", "Center", None, None, 3180, 3420, 3650),
        (98, "ชุดยูโร", "Center", None, None, 7770, 8480, 9070),
        (99, "มือเปิดฝากระโปรงท้าย (ทำสีตัวรถ)", "Center", None, 2120, None, None, None),
    ]

    # Populate MINI Cooper Sheet
    mini_row_count = 1
    mini_excel_row = 2

    for item_no, subject, pos, replace_p, mini_p, bmw1, bmw2, bmw3 in PDF_ITEMS_99:
        if mini_p is not None:
            row_vals = [
                mini_row_count,
                "MINI Cooper",
                "Cooper - R56",
                subject,
                pos,
                None, # Province (Blank)
                replace_p,
                mini_p, # S (Paint Price)
                mini_p, # M (Paint Price - Same)
                mini_p, # L (Paint Price - Same)
                REMARK_COLOR_PEARL
            ]
            ws_mini.append(row_vals)

            fill = zebra_fill if mini_excel_row % 2 == 0 else white_fill
            for col_idx in range(1, 12):
                cell = ws_mini.cell(row=mini_excel_row, column=col_idx)
                cell.fill = fill
                cell.font = data_font
                cell.border = thin_border

                if col_idx in [1, 2, 3, 5]:
                    cell.alignment = align_center
                    if col_idx == 3: cell.font = bold_font
                elif col_idx in [7, 8, 9, 10]:
                    cell.alignment = align_right
                    if cell.value is not None: cell.number_format = '#,##0'
                else:
                    cell.alignment = align_left

            ws_mini.row_dimensions[mini_excel_row].height = 22
            mini_row_count += 1
            mini_excel_row += 1

    # Populate BMW Sheet (Unpacked into 10 single models per item)
    bmw_row_count = 1
    bmw_excel_row = 2

    bmw_groups = [
        (BMW_G1, 5), # BMW1 Price
        (BMW_G2, 6), # BMW2 Price
        (BMW_G3, 7), # BMW3 Price
    ]

    for item_no, subject, pos, replace_p, mini_p, bmw1_p, bmw2_p, bmw3_p in PDF_ITEMS_99:
        bmw_prices = [bmw1_p, bmw2_p, bmw3_p]
        for g_idx, (models_list, p_idx) in enumerate(bmw_groups):
            p_val = bmw_prices[g_idx]
            if p_val is not None:
                for model_name in models_list:
                    row_vals = [
                        bmw_row_count,
                        "BMW",
                        model_name,
                        subject,
                        pos,
                        None, # Province (Blank)
                        replace_p,
                        p_val, # S (Paint Price)
                        p_val, # M (Paint Price - Same)
                        p_val, # L (Paint Price - Same)
                        REMARK_COLOR_PEARL
                    ]
                    ws_bmw.append(row_vals)

                    fill = zebra_fill if bmw_excel_row % 2 == 0 else white_fill
                    for col_idx in range(1, 12):
                        cell = ws_bmw.cell(row=bmw_excel_row, column=col_idx)
                        cell.fill = fill
                        cell.font = data_font
                        cell.border = thin_border

                        if col_idx in [1, 2, 3, 5]:
                            cell.alignment = align_center
                            if col_idx == 3: cell.font = bold_font
                        elif col_idx in [7, 8, 9, 10]:
                            cell.alignment = align_right
                            if cell.value is not None: cell.number_format = '#,##0'
                        else:
                            cell.alignment = align_left

                    ws_bmw.row_dimensions[bmw_excel_row].height = 22
                    bmw_row_count += 1
                    bmw_excel_row += 1

    # Auto adjust column widths for both sheets
    for ws in [ws_bmw, ws_mini]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        ws.column_dimensions['C'].width = 20 # Model
        ws.column_dimensions['D'].width = 38 # Subject
        ws.column_dimensions['E'].width = 16 # Position (L/R)
        ws.column_dimensions['F'].width = 20 # Province (Blank)
        ws.column_dimensions['K'].width = 32 # Remark

    wb.save(excel_path)
    print(f"SUCCESS: Updated BMW sheet ({bmw_row_count - 1} rows) and MINI Cooper sheet ({mini_row_count - 1} rows) with identical price in S, M, L columns at {excel_path}")

if __name__ == "__main__":
    populate_bmw_and_mini_all_severity_cols()
