import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def update_column_f_to_garage_province():
    excel_path = "/Users/arthur/Documents/Customer/H Tech/Labor Motor/Labor_Motor_Car_Brands.xlsx"
    
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found at", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path)

    # Styling Setup
    font_family = "TH Sarabun New"
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # #0071e3 Premium Blue
    header_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Update Column F in all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.cell(row=1, column=6, value="จังหวัดอู่ซ่อม")
        cell = ws.cell(row=1, column=6)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
        # Auto adjust column width
        ws.column_dimensions['F'].width = 22

    wb.save(excel_path)
    print("SUCCESS: Updated Column F header to 'จังหวัดอู่ซ่อม' across all sheets at", excel_path)

if __name__ == "__main__":
    update_column_f_to_garage_province()
